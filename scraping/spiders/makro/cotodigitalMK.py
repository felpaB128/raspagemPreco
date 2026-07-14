import csv
import json
import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import quote

import scrapy
from openpyxl import load_workbook


class CotodigitalMkSpider(scrapy.Spider):
    name = "cotodigital_mk"
    allowed_domains = ["ac.cnstrc.com"]

    custom_settings = {
        "ROBOTSTXT_OBEY": False,
        "DOWNLOAD_DELAY": 0.5,
        "CONCURRENT_REQUESTS_PER_DOMAIN": 2,
        "RETRY_TIMES": 2,
        "DOWNLOAD_TIMEOUT": 30,
        "FEED_EXPORT_ENCODING": "utf-8",
        "LOG_LEVEL": "INFO",
        "FEED_EXPORT_FIELDS": [
            "articulo_nr",
            "articulo_descripcion",
            "ean_entrada",
            "area",
            "marca_coto",
            "preco_por_coto",
            "preco_de_coto",
            "oferta_coto",
            "preco_referencia_coto",
            "desconto_percentual_coto",
            "tipo_oferta_coto",
            "sku_coto",
            "url_produto",
            "imagem",
            "search_url",
        ],
    }

    HEADER_ALIASES = {
        "articulo nr": "Artículo NR",
        "artículo nr": "Artículo NR",
        "codigo interno do concorrente": "Artículo NR",
        "cod interno do concorrente": "Artículo NR",
        "codigo interno": "Artículo NR",
        "cod interno": "Artículo NR",
        "articulo descripcion": "Artículo DESCRIPCION",
        "artículo descripcion": "Artículo DESCRIPCION",
        "descripcion": "Artículo DESCRIPCION",
        "descripción": "Artículo DESCRIPCION",
        "articulo": "Artículo DESCRIPCION",
        "ean": "EAN",
        "codigo de barras": "EAN",
        "cod barras": "EAN",
        "barcode": "EAN",
        "area": "AREA",
        "área": "AREA",
        "main group": "MAIN GROUP",
        "grupo principal": "MAIN GROUP",
        "grupo": "GRUPO",
        "ecompetidor": "eCompetidor",
        "competidor": "eCompetidor",
    }

    def __init__(self, input_file=None, store_id="200", sheet_name=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.input_file = input_file or "input.csv"
        self.store_id = str(store_id)
        self.sheet_name = sheet_name
        self.api_key = "key_r6xzz4IAoTWcipni"

    async def start(self):
        rows = self.filter_rows_by_competitor(
            self.load_input_rows(self.input_file, self.sheet_name),
            target="Coto Ciudadela",
        )

        headers = {
            "Accept": "application/json, text/plain, */*",
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/137.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "es-AR,es;q=0.9,en;q=0.8",
            "Referer": "https://www.cotodigital.com.ar/",
            "Origin": "https://www.cotodigital.com.ar",
        }

        for row in rows:
            descricao = self.clean_text(row.get("Artículo DESCRIPCION"))
            ean = self.clean_ean(row.get("EAN"))

            if ean:
                query = ean
                query_type = "ean"
            elif descricao:
                query = descricao
                query_type = "descripcion"
            else:
                yield self.build_empty_item(row, erro="Linha sem EAN e sem descrição.")
                continue

            yield scrapy.Request(
                url=self.build_search_url(query),
                callback=self.parse_search,
                errback=self.errback_search,
                dont_filter=True,
                headers=headers,
                meta={
                    "row": row,
                    "query": query,
                    "query_type": query_type,
                    "headers_used": headers,
                },
            )

    def build_search_url(self, query):
        q = quote(self.clean_text(query), safe="")
        return (
            "https://ac.cnstrc.com/search/"
            f"{q}"
            f"?c=cio-ui-autocomplete-1.29.3"
            f"&key={self.api_key}"
            f"&i=client-generated"
            f"&s=22"
            f"&num_results_per_page=24"
            f"&page=1"
            f"&fmt_options[hidden_fields]=price,discounts,product_brand,product_main_ean,sku_id,sku_plu,url,image_url,product_large_image_url,product_list_price,sale_type,groups,store_availability,sku_description,sku_display_name"
        )

    def parse_search(self, response):
        row = response.meta["row"]
        data = response.json().get("response", {}) or {}
        results = data.get("results", []) or []
        best = self.choose_best_result(row, results)

        if not best:
            yield self.build_empty_item(
                row=row,
                busca_tipo=response.meta.get("query_type", ""),
                busca_termo=response.meta.get("query", ""),
                search_url=response.url,
                total_resultados=data.get("total_num_results", 0),
                erro="Nenhum resultado compatível encontrado.",
            )
            return

        data_map = self.extract_result_fields(best)
        em_estoque = self.is_available_result(best)
        yield self.build_final_item_from_search(row, response, data_map, em_estoque)

    def errback_search(self, failure):
        req = getattr(failure, "request", None)
        row = req.meta.get("row", {}) if req else {}
        yield self.build_empty_item(
            row=row,
            busca_tipo=req.meta.get("query_type", "") if req else "",
            busca_termo=req.meta.get("query", "") if req else "",
            search_url=req.url if req else "",
            erro=repr(failure.value),
        )

    def choose_best_result(self, row, results):
        ean_entrada = self.clean_ean(row.get("EAN"))
        desc_entrada = self.normalize_text(row.get("Artículo DESCRIPCION"))
        categorias_entrada = " ".join(
            x for x in [
                self.normalize_text(row.get("AREA")),
                self.normalize_text(row.get("MAIN GROUP")),
                self.normalize_text(row.get("GRUPO")),
            ] if x
        )

        best = None
        best_score = -1

        for result in results:
            d = self.extract_result_fields(result)
            nome = self.normalize_text(d.get("nome"))
            marca = self.normalize_text(d.get("marca"))
            ean_result = self.clean_ean(d.get("ean"))
            sku_result = self.clean_text(d.get("sku"))

            group_names = []
            for g in (result.get("data", {}) or {}).get("groups", []) or []:
                group_names.append(self.normalize_text(g.get("display_name")))
            categoria_result = " ".join(x for x in group_names if x)

            score = 0

            if ean_entrada and ean_result and ean_entrada == ean_result:
                score += 100

            if ean_entrada and sku_result and ean_entrada in sku_result:
                score += 15

            if desc_entrada and nome:
                score += int(100 * SequenceMatcher(None, desc_entrada, nome).ratio())

            if categorias_entrada and categoria_result:
                score += int(30 * SequenceMatcher(None, categorias_entrada, categoria_result).ratio())

            if marca and desc_entrada and marca in desc_entrada:
                score += 10

            if score > best_score:
                best = result
                best_score = score

        return best

    def extract_result_fields(self, result):
        data = result.get("data", {}) or {}
        store_price = self._get_store_price_from_result(data, self.store_id)
        discount = self._get_first_discount(data)

        preco_por = ""
        preco_de = ""
        desconto_percentual = ""
        tipo_oferta = ""

        if discount:
            preco_por = discount.get("discountPrice") or ""
            preco_de = self._extract_number_from_text(discount.get("regularPriceText") or "")
            tipo_oferta = discount.get("discountText") or ""
            desconto_percentual = self.extract_only_percentage(tipo_oferta)

        if not preco_por and store_price:
            preco_por = store_price.get("formatPrice") or ""

        if not preco_de and store_price:
            preco_de = store_price.get("listPrice") or ""

        if not tipo_oferta:
            sale_type = data.get("sale_type") or []
            if isinstance(sale_type, list):
                tipo_oferta = " ".join(str(x) for x in sale_type if x)
            else:
                tipo_oferta = str(sale_type or "")
            if not desconto_percentual:
                desconto_percentual = self.extract_only_percentage(tipo_oferta)

        imagem = (
            data.get("image_url")
            or data.get("product_large_image_url")
            or data.get("product_medium_image_url")
            or ""
        )

        return {
            "nome": data.get("sku_display_name") or data.get("sku_description") or result.get("value", "") or "",
            "marca": data.get("product_brand") or "",
            "preco": preco_por,
            "preco_referencia": preco_de,
            "desconto_percentual": desconto_percentual,
            "tipo_oferta": tipo_oferta,
            "ean": data.get("product_main_ean") or "",
            "sku": data.get("sku_id") or data.get("sku_plu") or "",
            "url": data.get("url") or "",
            "imagem": imagem,
        }

    def build_final_item_from_search(self, row, response, data_map, em_estoque=True):
        preco_destaque = self._format_output_price(data_map.get("preco"))
        preco_ref = self._format_output_price(data_map.get("preco_referencia"))
        desconto = data_map.get("desconto_percentual")
        tipo_oferta = self.extract_only_percentage(data_map.get("tipo_oferta"))

        if not tipo_oferta and desconto not in (None, ""):
            tipo_oferta = f"{str(desconto).replace('.0', '')}%"

        if not em_estoque:
            preco_destaque = ""
            preco_ref = ""
            desconto = ""
            tipo_oferta = ""

        oferta_coto = "x" if tipo_oferta else ""
        preco_por_coto = preco_destaque if tipo_oferta else ""

        return {
            "articulo_nr": row.get("Artículo NR", ""),
            "articulo_descripcion": row.get("Artículo DESCRIPCION", ""),
            "ean_entrada": row.get("EAN", ""),
            "area": row.get("AREA", ""),
            "marca_coto": data_map.get("marca", ""),
            "preco_por_coto": preco_por_coto,
            "preco_de_coto": preco_ref,
            "oferta_coto": oferta_coto,
            "preco_referencia_coto": preco_ref,
            "desconto_percentual_coto": desconto if desconto not in (None, "") else "",
            "tipo_oferta_coto": tipo_oferta,
            "sku_coto": data_map.get("sku", ""),
            "url_produto": self.make_absolute_product_url(data_map.get("url", "")),
            "imagem": data_map.get("imagem", ""),
            "search_url": getattr(response, "url", ""),
        }

    def is_available_result(self, result):
        data = result.get("data", {}) or {}
        stores = data.get("store_availability") or []
        target = self.store_id.zfill(3)

        if isinstance(stores, list):
            return target in [str(x).zfill(3) for x in stores]

        return True

    def _get_store_price_from_result(self, data, store_id):
        prices = data.get("price") or []
        target = str(store_id).zfill(3)

        for p in prices:
            if str(p.get("store", "")).zfill(3) == target:
                return p

        return prices[0] if prices else {}

    def _get_first_discount(self, data):
        discounts = data.get("discounts") or []
        return discounts[0] if discounts else {}

    def _extract_number_from_text(self, text):
        text = str(text or "")
        m = re.search(r'(\d+(?:[.,]\d{1,2})?)', text.replace(".", "").replace(",", "."))
        return m.group(1) if m else ""

    def _format_output_price(self, value):
        if value in (None, "", [], {}):
            return ""
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        return str(value).strip()

    def extract_only_percentage(self, value):
        texto = self.clean_text(value)
        if not texto:
            return ""

        m = re.search(r"(\d+(?:[.,]\d+)?)\s*%", texto)
        if m:
            numero = m.group(1).replace(",", ".")
            if numero.endswith(".0"):
                numero = numero[:-2]
            return f"{numero}%"

        m = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:dto|descuento)", texto, re.I)
        if m:
            numero = m.group(1).replace(",", ".")
            if numero.endswith(".0"):
                numero = numero[:-2]
            return f"{numero}%"

        return ""

    def make_absolute_product_url(self, url):
        url = self.clean_text(url)
        if not url:
            return ""
        if url.startswith(("http://", "https://")):
            return url
        if url.startswith("_/"):
            return f"https://www.cotodigital.com.ar/sitios/cdigi/productos/{url}"
        return f"https://www.cotodigital.com.ar{url}"
    
    def build_empty_item(self, row, busca_tipo="", busca_termo="", search_url="", total_resultados="", erro=""):
        return {
            "articulo_nr": row.get("Artículo NR", ""),
            "articulo_descripcion": row.get("Artículo DESCRIPCION", ""),
            "ean_entrada": row.get("EAN", ""),
            "area": row.get("AREA", ""),
            "marca_coto": "",
            "preco_por_coto": "",
            "preco_de_coto": "",
            "oferta_coto": "",
            "preco_referencia_coto": "",
            "desconto_percentual_coto": "",
            "tipo_oferta_coto": "",
            "sku_coto": "",
            "url_produto": "",
            "imagem": "",
            "search_url": search_url,
        }

    def load_input_rows(self, file_path, sheet_name=None):
        path = self.resolve_input_path(file_path)

        if path.suffix.lower() == ".csv":
            with open(path, "r", encoding="utf-8-sig", newline="") as f:
                return self.rows_from_matrix(list(csv.reader(f)))

        if path.suffix.lower() in [".xlsx", ".xlsm"]:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            try:
                ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else self.choose_best_sheet(wb)
                rows = [[self.clean_cell_value(c) for c in row] for row in ws.iter_rows(values_only=True)]
                return self.rows_from_matrix(rows)
            finally:
                wb.close()

        raise ValueError(f"Extensão não suportada: {path.suffix.lower()}")

    def resolve_input_path(self, file_path):
        path = Path(file_path)

        if not path.is_absolute():
            candidates = [
                Path.cwd() / path,
                Path(__file__).resolve().parent / path,
                Path(__file__).resolve().parent.parent / path,
                Path(__file__).resolve().parent.parent.parent / path,
            ]
            for candidate in candidates:
                if candidate.exists():
                    return candidate

        if not path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path} | cwd={Path.cwd()}")

        return path

    def rows_from_matrix(self, all_rows):
        if not all_rows:
            return []

        header_idx = self.find_header_row_index(all_rows)
        headers = [self.map_header_name(h) for h in all_rows[header_idx]]
        out = []

        for raw in all_rows[header_idx + 1:]:
            row = {
                h: self.clean_cell_value(raw[i] if i < len(raw) else "")
                for i, h in enumerate(headers)
                if h
            }
            if any(self.clean_text(v) for v in row.values()):
                out.append(row)

        return out

    def choose_best_sheet(self, workbook):
        best_ws = workbook[workbook.sheetnames[0]]
        best_score = -1

        for name in workbook.sheetnames:
            ws = workbook[name]
            preview = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 15:
                    break
                preview.append([self.clean_cell_value(c) for c in row])

            score = self.score_sheet(preview)
            if score > best_score:
                best_ws = ws
                best_score = score

        return best_ws

    def score_sheet(self, rows):
        best = 0
        for row in rows[:15]:
            normalized = [self.normalize_header(v) for v in row if self.clean_text(v)]
            best = max(best, sum(1 for col in normalized if col in self.HEADER_ALIASES))
        return best

    def find_header_row_index(self, rows):
        best_idx = 0
        best_score = -1

        for idx in range(min(len(rows), 20)):
            normalized = [self.normalize_header(v) for v in rows[idx] if self.clean_text(v)]
            found = {self.HEADER_ALIASES[col] for col in normalized if col in self.HEADER_ALIASES}
            score = len(found) + (2 if "EAN" in found else 0) + (2 if "Artículo DESCRIPCION" in found else 0)
            if score > best_score:
                best_idx = idx
                best_score = score

        return best_idx

    def map_header_name(self, header):
        return self.HEADER_ALIASES.get(self.normalize_header(header), self.clean_text(header))

    def filter_rows_by_competitor(self, rows, target="Coto Ciudadela"):
        target_norm = self.normalize_text(target)
        return [r for r in rows if self.normalize_text(r.get("eCompetidor")) == target_norm]

    def clean_cell_value(self, value):
        if value is None:
            return ""
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value)
        return str(value).strip()

    def clean_text(self, value):
        return "" if value is None else str(value).strip()

    def clean_ean(self, value):
        return "" if value is None else re.sub(r"\D+", "", str(value).strip())

    def normalize_header(self, value):
        value = self.clean_text(value).replace("\n", " ").replace("\r", " ").replace("_", " ")
        value = re.sub(r"\s+", " ", value).strip().lower()
        value = unicodedata.normalize("NFKD", value)
        return "".join(ch for ch in value if not unicodedata.combining(ch))

    def normalize_text(self, value):
        value = self.clean_text(value).lower()
        value = unicodedata.normalize("NFKD", value)
        value = "".join(ch for ch in value if not unicodedata.combining(ch))
        value = re.sub(r"[^a-z0-9\s]", " ", value, flags=re.IGNORECASE)
        return re.sub(r"\s+", " ", value).strip()