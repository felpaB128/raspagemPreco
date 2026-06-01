import csv
import re
import json
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import quote

from scrapy import Spider, Request, Selector

try:
    import openpyxl
except ImportError:
    openpyxl = None


class SupermercadosDiaMKSpider(Spider):
    name = "supermercadosdia_mk"
    allowed_domains = ["diaonline.supermercadosdia.com.ar", "supermercadosdia.com.ar"]

    custom_settings = {
        "ZYTE_API_TRANSPARENT_MODE": True,
        "CONCURRENT_REQUESTS_PER_DOMAIN": 4,
        "DOWNLOAD_DELAY": 0.2,
        "LOG_LEVEL": "INFO",
        "FEED_EXPORT_ENCODING": "utf-8",
        "FEEDS": {
            "dia_ar.csv": {
                "format": "csv",
                "encoding": "utf-8",
                "fields": [
                    "loja",
                    "tipo_busca",
                    "busca_valor",
                    "ean",
                    "nome",
                    "marca",
                    "preco",
                    "precoPor",
                    "precoDe",
                    "oferta",
                    "link",
                    "status_busca",
                ],
            }
        },
    }

    API_EAN = "https://diaonline.supermercadosdia.com.ar/api/catalog_system/pub/products/search?fq=alternateIds_Ean:{}"

    def __init__(self, ean=None, arquivo_entrada=None, termo=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if not ean and not arquivo_entrada and not termo:
            raise ValueError(
                "Passe ean, termo ou arquivo_entrada. "
                "Ex.: -a ean=7790272001005 | -a termo='coca cola' | -a arquivo_entrada=arquivo.xlsx"
            )

        self.ean = ean
        self.termo = termo
        self.arquivo_entrada = arquivo_entrada

    # -----------------------------
    # leitura de arquivo
    # -----------------------------
    def _resolver_caminho_arquivo(self, caminho_str: str) -> Path:
        caminho = Path(caminho_str).expanduser()

        if caminho.is_absolute():
            return caminho

        candidatos = [
            Path.cwd() / caminho,
            Path(__file__).resolve().parents[4] / caminho,
            Path(__file__).resolve().parents[3] / caminho,
        ]

        for candidato in candidatos:
            if candidato.exists():
                return candidato

        return candidatos[0]

    def _eh_competidor_dia(self, valor: str) -> bool:
        if not valor:
            return False
        return "dia" in valor.strip().lower()

    def _normalizar_texto(self, texto):
        if texto is None:
            return ""
        texto = str(texto).strip().lower()
        texto = re.sub(r"\s+", " ", texto)
        return texto

    def _campos_header(self, fieldnames):
        return {str(c).lower().strip(): c for c in fieldnames if c is not None}

    def _descobrir_coluna(self, nomes, candidatos):
        for c in candidatos:
            if c in nomes:
                return nomes[c]
        return None

    def _ler_registros_csv(self, caminho: Path):
        registros = []

        with caminho.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("CSV sem cabeçalho.")

            nomes = self._campos_header(reader.fieldnames)

            coluna_ean = self._descobrir_coluna(nomes, [
                "ean", "código ean", "codigo ean", "codigo_ean",
                "codigoean", "ean 13", "cod ean", "cod_ean"
            ])

            coluna_nome = self._descobrir_coluna(nomes, [
                "nome", "produto", "descricao", "descrição",
                "descricao produto", "descrição produto", "product_name", "nome produto"
            ])

            coluna_competidor = self._descobrir_coluna(nomes, [
                "competidor", "competidor ", "concorrente"
            ])

            if not coluna_ean and not coluna_nome:
                raise ValueError(
                    f"Não encontrei coluna de EAN nem de nome no CSV. Cabeçalho: {reader.fieldnames}"
                )

            for row in reader:
                if coluna_competidor:
                    valor_comp = row.get(coluna_competidor) or ""
                    if not self._eh_competidor_dia(valor_comp):
                        continue

                ean = str(row.get(coluna_ean) or "").strip() if coluna_ean else ""
                nome = str(row.get(coluna_nome) or "").strip() if coluna_nome else ""

                if ean or nome:
                    registros.append({
                        "ean": ean or None,
                        "nome": nome or None,
                    })

        unicos = []
        vistos = set()
        for r in registros:
            chave = (r.get("ean") or "", self._normalizar_texto(r.get("nome")))
            if chave not in vistos:
                vistos.add(chave)
                unicos.append(r)

        return unicos

    def _ler_registros_xlsx(self, caminho: Path):
        if openpyxl is None:
            raise RuntimeError("openpyxl não está instalado. Rode: pip install openpyxl")

        wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Planilha vazia.")

        header_row_idx = None
        header = None

        for idx, row in enumerate(rows[:20]):
            nomes_linha = [str(h).strip().lower() if h is not None else "" for h in row]
            if any(
                n in (
                    "ean", "código ean", "codigo ean", "codigo_ean",
                    "codigoean", "ean 13", "cod ean", "cod_ean",
                    "nome", "produto", "descricao", "descrição", "nome produto"
                )
                for n in nomes_linha
            ):
                header_row_idx = idx
                header = [str(h).strip() if h is not None else "" for h in row]
                break

        if header_row_idx is None:
            for idx, row in enumerate(rows):
                if any(c is not None and str(c).strip() for c in row):
                    header_row_idx = idx
                    header = [str(h).strip() if h is not None else "" for h in row]
                    break

        if header_row_idx is None or header is None:
            raise ValueError("Não encontrei nenhuma linha de cabeçalho na planilha.")

        header_norm = [h.lower().strip() for h in header]
        nomes = {c: idx for idx, c in enumerate(header_norm)}

        idx_ean = None
        for chave in (
            "ean", "código ean", "codigo ean", "codigo_ean",
            "codigoean", "ean 13", "cod ean", "cod_ean"
        ):
            if chave in nomes:
                idx_ean = nomes[chave]
                break

        idx_nome = None
        for chave in (
            "nome", "produto", "descricao", "descrição",
            "descricao produto", "descrição produto", "product_name", "nome produto"
        ):
            if chave in nomes:
                idx_nome = nomes[chave]
                break

        idx_competidor = None
        for chave in ("competidor", "concorrente"):
            if chave in nomes:
                idx_competidor = nomes[chave]
                break

        if idx_ean is None and idx_nome is None:
            raise ValueError(
                f"Não encontrei coluna de EAN nem de nome no XLSX. Cabeçalho detectado: {header}"
            )

        registros = []
        for row in rows[header_row_idx + 1:]:
            if not row:
                continue

            if idx_competidor is not None and idx_competidor < len(row):
                valor_comp = row[idx_competidor]
                if not self._eh_competidor_dia(str(valor_comp) if valor_comp is not None else ""):
                    continue

            ean = None
            nome = None

            if idx_ean is not None and idx_ean < len(row):
                valor = row[idx_ean]
                if valor is not None:
                    ean = str(valor).strip() or None

            if idx_nome is not None and idx_nome < len(row):
                valor = row[idx_nome]
                if valor is not None:
                    nome = str(valor).strip() or None

            if ean or nome:
                registros.append({
                    "ean": ean,
                    "nome": nome,
                })

        unicos = []
        vistos = set()
        for r in registros:
            chave = (r.get("ean") or "", self._normalizar_texto(r.get("nome")))
            if chave not in vistos:
                vistos.add(chave)
                unicos.append(r)

        return unicos

    def _ler_registros_arquivo(self, caminho_str: str):
        caminho = self._resolver_caminho_arquivo(caminho_str)

        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

        if caminho.suffix.lower() == ".csv":
            return self._ler_registros_csv(caminho)
        elif caminho.suffix.lower() == ".xlsx":
            return self._ler_registros_xlsx(caminho)

        raise ValueError("Arquivo de entrada deve ser .csv ou .xlsx")

    # -----------------------------
    # util
    # -----------------------------
    def montar_url_busca(self, termo: str) -> str:
        termo = quote(str(termo).strip())
        return f"https://diaonline.supermercadosdia.com.ar/{termo}?_q={termo}&map=ft"

    def montar_url_api_ean(self, ean: str) -> str:
        return self.API_EAN.format(quote(str(ean).strip(), safe=":"))

    def _get_selector(self, response):
        raw = getattr(response, "raw_api_response", None) or {}
        browser_html = raw.get("browserHtml")
        if browser_html:
            return Selector(text=browser_html)
        return response

    def _to_float(self, valor):
        if valor is None:
            return None

        if isinstance(valor, (int, float)):
            return float(valor)

        texto = str(valor).strip()
        if not texto:
            return None

        texto = texto.replace("$", "").replace("\xa0", " ")
        texto = texto.replace(".", "").replace(",", ".")
        texto = re.sub(r"[^\d.]", "", texto)

        try:
            return float(texto) if texto else None
        except Exception:
            return None

    def _price_to_str(self, valor):
        if valor is None:
            return None
        valor = float(valor)
        if valor.is_integer():
            return str(int(valor))
        return f"{valor:.2f}"

    def normalizar_precos(self, preco=None, preco_por=None, preco_de=None):
        preco_f = self._to_float(preco)
        preco_por_f = self._to_float(preco_por)
        preco_de_f = self._to_float(preco_de)

        if preco_por_f is None and preco_f is not None:
            preco_por_f = preco_f

        oferta = None

        if preco_por_f is not None and preco_de_f is not None and preco_de_f > preco_por_f:
            preco_final = preco_por_f
            oferta = "x"
        else:
            preco_final = preco_por_f if preco_por_f is not None else preco_de_f
            preco_por_f = None
            preco_de_f = None

        return {
            "preco": self._price_to_str(preco_final),
            "precoPor": self._price_to_str(preco_por_f),
            "precoDe": self._price_to_str(preco_de_f),
            "oferta": oferta,
        }

    # -----------------------------
    # NOVO: origem estruturada de preço
    # -----------------------------
    def _buscar_jsons_na_pagina(self, sel: Selector):
        jsons = []

        scripts = sel.css("script::text").getall()
        for s in scripts:
            if not s:
                continue
            if ("Price" in s) or ("commertialOffer" in s) or ("ListPrice" in s) or ('"price"' in s) or ('"offers"' in s):
                jsons.append(s)

        return jsons

    def _extrair_precos_de_obj(self, obj):
        if isinstance(obj, dict):
            if "commertialOffer" in obj and isinstance(obj["commertialOffer"], dict):
                offer = obj["commertialOffer"]
                preco = offer.get("Price")
                preco_de = offer.get("ListPrice")
                if preco is not None or preco_de is not None:
                    return self.normalizar_precos(preco=preco, preco_por=preco, preco_de=preco_de)

            if "price" in obj and ("listPrice" in obj or "Price" in obj or "ListPrice" in obj):
                preco = obj.get("price") or obj.get("Price")
                preco_de = obj.get("listPrice") or obj.get("ListPrice")
                if preco is not None or preco_de is not None:
                    return self.normalizar_precos(preco=preco, preco_por=preco, preco_de=preco_de)

            if "offers" in obj:
                offers = obj.get("offers")
                if isinstance(offers, dict):
                    preco = offers.get("price") or offers.get("lowPrice")
                    preco_de = offers.get("highPrice") or offers.get("listPrice")
                    if preco is not None or preco_de is not None:
                        return self.normalizar_precos(preco=preco, preco_por=preco, preco_de=preco_de)
                elif isinstance(offers, list):
                    for off in offers:
                        achou = self._extrair_precos_de_obj(off)
                        if achou and achou.get("preco"):
                            return achou

            for v in obj.values():
                achou = self._extrair_precos_de_obj(v)
                if achou and achou.get("preco"):
                    return achou

        elif isinstance(obj, list):
            for item in obj:
                achou = self._extrair_precos_de_obj(item)
                if achou and achou.get("preco"):
                    return achou

        return None

    def _extrair_precos_json_pagina(self, sel: Selector):
        for bloco in self._buscar_jsons_na_pagina(sel):
            texto = (bloco or "").strip()
            if not texto:
                continue

            try:
                obj = json.loads(texto)
                achou = self._extrair_precos_de_obj(obj)
                if achou and achou.get("preco"):
                    return achou
            except Exception:
                pass

            candidatos = []

            if "commertialOffer" in texto:
                candidatos.extend(re.findall(r'\{.*?"commertialOffer".*?\}\s*(?=,?\s*[\]\}])', texto, flags=re.DOTALL))

            if '"offers"' in texto or '"price"' in texto:
                candidatos.extend(re.findall(r'\{.*?"offers".*?\}\s*(?=,?\s*[\]\}])', texto, flags=re.DOTALL))

            for trecho in candidatos:
                try:
                    obj = json.loads(trecho)
                    achou = self._extrair_precos_de_obj(obj)
                    if achou and achou.get("preco"):
                        return achou
                except Exception:
                    continue

        return None

    def extrair_preco_regex(self, texto: str):
        if not texto:
            return None

        texto = " ".join(texto.split())
        padrao = r"\$\s*\d{1,3}(?:\.\d{3})*(?:,\d{2})?"

        m = re.search(padrao, texto, flags=re.IGNORECASE)
        if m:
            return m.group(0)
        return None

    def extrair_todos_precos_regex(self, texto: str):
        if not texto:
            return []

        texto = " ".join(texto.split())
        padrao = r"\$\s*\d{1,3}(?:\.\d{3})*(?:,\d{2})?"
        return re.findall(padrao, texto, flags=re.IGNORECASE)

    def extrair_precos_pdp(self, sel: Selector):
        precos_json = self._extrair_precos_json_pagina(sel)
        if precos_json and precos_json.get("preco"):
            return precos_json

        preco_por = None
        preco_de = None

        seletores_preco_por = [
            ".vtex-product-price-1-x-sellingPriceValue::text",
            ".vtex-product-price-1-x-currencyContainer .vtex-product-price-1-x-sellingPriceValue::text",
            "[class*='sellingPriceValue']::text",
            "[data-testid='price']::text",
        ]

        seletores_preco_de = [
            ".vtex-product-price-1-x-listPriceValue::text",
            "[class*='listPriceValue']::text",
            "[class*='listPrice']::text",
        ]

        for s in seletores_preco_por:
            textos = sel.css(s).getall()
            textos = [" ".join(t.split()) for t in textos if t and t.strip()]
            for t in textos:
                preco = self.extrair_preco_regex(t)
                if preco:
                    preco_por = preco
                    break
            if preco_por:
                break

        for s in seletores_preco_de:
            textos = sel.css(s).getall()
            textos = [" ".join(t.split()) for t in textos if t and t.strip()]
            for t in textos:
                preco = self.extrair_preco_regex(t)
                if preco:
                    preco_de = preco
                    break
            if preco_de:
                break

        if not preco_por:
            textos = [t.strip() for t in sel.css("body ::text").getall() if t and t.strip()]
            bloco = " ".join(textos)
            precos = self.extrair_todos_precos_regex(bloco)
            if precos:
                preco_por = precos[0]
                if len(precos) > 1:
                    preco_de = preco_de or precos[1]

        return self.normalizar_precos(preco=preco_por, preco_por=preco_por, preco_de=preco_de)

    def extrair_precos_card(self, card: Selector):
        precos_json = self._extrair_precos_json_pagina(card)
        if precos_json and precos_json.get("preco"):
            return precos_json

        preco_por = None
        preco_de = None

        seletores_preco_por = [
            ".vtex-product-price-1-x-sellingPriceValue::text",
            ".vtex-product-price-1-x-currencyContainer .vtex-product-price-1-x-sellingPriceValue::text",
            "[class*='sellingPriceValue']::text",
            "[data-testid='price']::text",
            ".vtex-product-price-1-x-currencyContainer::text",
            "[class*='currencyContainer']::text",
        ]

        seletores_preco_de = [
            ".vtex-product-price-1-x-listPriceValue::text",
            "[class*='listPriceValue']::text",
            "[class*='listPrice']::text",
        ]

        for s in seletores_preco_por:
            textos = card.css(s).getall()
            textos = [" ".join(t.split()) for t in textos if t and t.strip()]
            for t in textos:
                preco = self.extrair_preco_regex(t)
                if preco:
                    preco_por = preco
                    break
            if preco_por:
                break

        for s in seletores_preco_de:
            textos = card.css(s).getall()
            textos = [" ".join(t.split()) for t in textos if t and t.strip()]
            for t in textos:
                preco = self.extrair_preco_regex(t)
                if preco:
                    preco_de = preco
                    break
            if preco_de:
                break

        if not preco_por:
            textos = [t.strip() for t in card.css("::text").getall() if t and t.strip()]
            bloco = " ".join(textos)
            precos = self.extrair_todos_precos_regex(bloco)
            if precos:
                preco_por = precos[0]
                if len(precos) > 1:
                    preco_de = preco_de or precos[1]

        return self.normalizar_precos(preco=preco_por, preco_por=preco_por, preco_de=preco_de)

    def extrair_marca(self, nome):
        if not nome:
            return None
        partes = nome.split()
        return partes[0] if partes else None

    def extrair_nome_listagem(self, card):
        seletores = [
            "h2::text",
            "h3::text",
            ".vtex-product-summary-2-x-productBrand::text",
            ".vtex-product-summary-2-x-productName::text",
            "a::text",
            "span::text",
        ]
        for sel in seletores:
            textos = card.css(sel).getall()
            textos = [" ".join(t.split()) for t in textos if t and t.strip()]
            textos = [
                t for t in textos
                if len(t) > 2 and "A un clic de llevarte el producto" not in t
            ]
            if textos:
                return textos[0]
        return None

    def _score_nome(self, nome_busca, nome_candidato):
        a = self._normalizar_texto(nome_busca)
        b = self._normalizar_texto(nome_candidato)
        if not a or not b:
            return 0

        ratio = SequenceMatcher(None, a, b).ratio()

        tokens_a = set(a.split())
        tokens_b = set(b.split())
        inter = len(tokens_a & tokens_b)
        bonus = inter / max(len(tokens_a), 1)

        return ratio * 0.7 + bonus * 0.3

    def _escolher_item_por_ean(self, produto, ean_busca):
        ean_busca = str(ean_busca).strip()
        itens = produto.get("items") or []

        for item in itens:
            if str(item.get("ean") or "").strip() == ean_busca:
                return item

        return itens[0] if itens else None

    def _escolher_seller(self, item):
        sellers = item.get("sellers") or []
        if not sellers:
            return None

        for seller in sellers:
            if seller.get("sellerDefault"):
                return seller

        for seller in sellers:
            offer = seller.get("commertialOffer") or {}
            if offer.get("AvailableQuantity", 0) > 0:
                return seller

        return sellers[0]

    def _montar_item_api(self, produto, ean_busca):
        item = self._escolher_item_por_ean(produto, ean_busca)
        if not item:
            return None

        seller = self._escolher_seller(item)
        offer = (seller or {}).get("commertialOffer") or {}

        preco = offer.get("Price")
        preco_de = offer.get("ListPrice")
        precos = self.normalizar_precos(preco=preco, preco_por=preco, preco_de=preco_de)

        link = produto.get("link")
        if link and link.startswith("/"):
            link = "https://diaonline.supermercadosdia.com.ar" + link

        nome = produto.get("productName") or item.get("nameComplete") or item.get("name")
        marca = produto.get("brand") or self.extrair_marca(nome)

        return {
            "loja": "dia_ar",
            "tipo_busca": "ean",
            "busca_valor": ean_busca,
            "ean": ean_busca,
            "nome": nome,
            "marca": marca,
            "preco": precos.get("preco"),
            "precoPor": precos.get("precoPor"),
            "precoDe": precos.get("precoDe"),
            "oferta": precos.get("oferta"),
            "link": link,
            "status_busca": "encontrado_via_api" if precos.get("preco") else "encontrado_via_api_sem_preco",
        }

    def _item_vazio(self, tipo_busca, busca_valor, ean=None, nome=None, link=None, status="nao_encontrado"):
        return {
            "loja": "dia_ar",
            "tipo_busca": tipo_busca,
            "busca_valor": busca_valor,
            "ean": ean,
            "nome": nome,
            "marca": None,
            "preco": None,
            "precoPor": None,
            "precoDe": None,
            "oferta": None,
            "link": link,
            "status_busca": status,
        }

    # -----------------------------
    # start
    # -----------------------------
    async def start(self):
        if self.arquivo_entrada:
            caminho_resolvido = self._resolver_caminho_arquivo(self.arquivo_entrada)
            self.logger.info("Lendo arquivo de entrada: %s", caminho_resolvido)

            registros = self._ler_registros_arquivo(self.arquivo_entrada)
            self.logger.info("Processando %d registros do arquivo", len(registros))

            for reg in registros:
                ean = reg.get("ean")
                nome = reg.get("nome")

                if ean:
                    yield Request(
                        url=self.montar_url_api_ean(ean),
                        callback=self.parse_api_ean,
                        dont_filter=True,
                        meta={
                            "busca_valor": str(ean),
                            "tipo_busca": "ean",
                            "ean_entrada": ean,
                            "nome_entrada": nome,
                        },
                    )
                elif nome:
                    yield Request(
                        url=self.montar_url_busca(nome),
                        callback=self.parse_search,
                        dont_filter=True,
                        meta={
                            "busca_valor": str(nome),
                            "tipo_busca": "nome",
                            "ean_entrada": ean,
                            "nome_entrada": nome,
                            "zyte_api_automap": {
                                "browserHtml": True,
                                "actions": [{"action": "scrollBottom"}],
                            },
                        },
                    )
            return

        valor_busca = self.termo or self.ean
        tipo_busca = "termo" if self.termo else "ean"

        if tipo_busca == "ean":
            yield Request(
                url=self.montar_url_api_ean(valor_busca),
                callback=self.parse_api_ean,
                dont_filter=True,
                meta={
                    "busca_valor": str(valor_busca),
                    "tipo_busca": "ean",
                    "ean_entrada": str(valor_busca),
                    "nome_entrada": None,
                },
            )
        else:
            yield Request(
                url=self.montar_url_busca(valor_busca),
                callback=self.parse_search,
                dont_filter=True,
                meta={
                    "busca_valor": str(valor_busca),
                    "tipo_busca": "termo",
                    "ean_entrada": None,
                    "nome_entrada": str(valor_busca),
                    "zyte_api_automap": {
                        "browserHtml": True,
                        "actions": [{"action": "scrollBottom"}],
                    },
                },
            )

    # -----------------------------
    # parse API EAN
    # -----------------------------
    def parse_api_ean(self, response):
        ean = response.meta.get("ean_entrada")
        nome_entrada = response.meta.get("nome_entrada")

        try:
            data = response.json()
        except Exception:
            data = None

        if isinstance(data, list) and data:
            encontrou = False

            for produto in data:
                item = self._montar_item_api(produto, ean)
                if item:
                    encontrou = True
                    yield item

            if encontrou:
                return

        if nome_entrada:
            yield Request(
                url=self.montar_url_busca(nome_entrada),
                callback=self.parse_search,
                dont_filter=True,
                meta={
                    "busca_valor": nome_entrada,
                    "tipo_busca": "nome",
                    "ean_entrada": ean,
                    "nome_entrada": nome_entrada,
                    "origem_fallback": "api_sem_match",
                    "zyte_api_automap": {
                        "browserHtml": True,
                        "actions": [{"action": "scrollBottom"}],
                    },
                },
            )
            return

        yield Request(
            url=self.montar_url_busca(ean),
            callback=self.parse_search,
            dont_filter=True,
            meta={
                "busca_valor": ean,
                "tipo_busca": "ean",
                "ean_entrada": ean,
                "nome_entrada": nome_entrada,
                "origem_fallback": "api_vazia",
                "zyte_api_automap": {
                    "browserHtml": True,
                    "actions": [{"action": "scrollBottom"}],
                },
            },
        )

    # -----------------------------
    # parse busca
    # -----------------------------
    def parse_search(self, response):
        busca_valor = response.meta.get("busca_valor")
        tipo_busca = response.meta.get("tipo_busca")
        ean_entrada = response.meta.get("ean_entrada")
        nome_entrada = response.meta.get("nome_entrada")
        sel = self._get_selector(response)

        self.logger.info("Busca %s | valor=%s | URL=%s", tipo_busca, busca_valor, response.url)

        links = sel.css("a::attr(href)").getall()
        links_produto = []

        for href in links:
            if not href:
                continue
            href = href.strip()

            if href.endswith("/p") or "/p?" in href or "/product/" in href.lower():
                abs_url = response.urljoin(href)
                if abs_url not in links_produto:
                    links_produto.append(abs_url)

        cards = sel.css("article, section, div")
        candidatos = []

        for card in cards:
            html_card = card.get() or ""
            if "$" not in html_card and "Price" not in html_card and "commertialOffer" not in html_card:
                continue

            nome = self.extrair_nome_listagem(card)
            precos = self.extrair_precos_card(card)

            if nome or precos.get("preco"):
                score = self._score_nome(nome_entrada or busca_valor, nome) if (nome_entrada or busca_valor) else 0
                candidatos.append({
                    "nome": nome,
                    "marca": self.extrair_marca(nome),
                    "preco": precos.get("preco"),
                    "precoPor": precos.get("precoPor"),
                    "precoDe": precos.get("precoDe"),
                    "oferta": precos.get("oferta"),
                    "link": response.url,
                    "score": score,
                })

        if candidatos and tipo_busca in ("nome", "termo"):
            melhor = sorted(candidatos, key=lambda x: (x["score"], x["preco"] is not None), reverse=True)[0]
            yield {
                "loja": "dia_ar",
                "tipo_busca": tipo_busca,
                "busca_valor": busca_valor,
                "ean": ean_entrada if tipo_busca == "nome" else None,
                "nome": melhor.get("nome"),
                "marca": melhor.get("marca"),
                "preco": melhor.get("preco"),
                "precoPor": melhor.get("precoPor"),
                "precoDe": melhor.get("precoDe"),
                "oferta": melhor.get("oferta"),
                "link": melhor.get("link"),
                "status_busca": "resultado_nome_em_listagem",
            }
            return

        if links_produto:
            for link in links_produto:
                yield Request(
                    url=link,
                    callback=self.parse_produto,
                    dont_filter=True,
                    meta={
                        "busca_valor": busca_valor,
                        "tipo_busca": tipo_busca,
                        "ean_entrada": ean_entrada,
                        "nome_entrada": nome_entrada,
                        "link_produto": link,
                        "zyte_api_automap": {
                            "browserHtml": True,
                        },
                    },
                )
            return

        if candidatos:
            for cand in candidatos[:1]:
                status = (
                    "resultado_em_listagem_com_preco"
                    if cand.get("preco")
                    else "resultado_em_listagem_sem_preco"
                )
                yield {
                    "loja": "dia_ar",
                    "tipo_busca": tipo_busca,
                    "busca_valor": busca_valor,
                    "ean": ean_entrada if tipo_busca in ("ean", "nome") else None,
                    "nome": cand.get("nome"),
                    "marca": cand.get("marca"),
                    "preco": cand.get("preco"),
                    "precoPor": cand.get("precoPor"),
                    "precoDe": cand.get("precoDe"),
                    "oferta": cand.get("oferta"),
                    "link": cand.get("link"),
                    "status_busca": status,
                }
            return

        yield self._item_vazio(
            tipo_busca=tipo_busca,
            busca_valor=busca_valor,
            ean=ean_entrada if tipo_busca in ("ean", "nome") else None,
            nome=nome_entrada if tipo_busca == "nome" else None,
            link=response.url,
            status="nao_indexado_na_busca",
        )

    # -----------------------------
    # parse produto (PDP)
    # -----------------------------
    def parse_produto(self, response):
        busca_valor = response.meta.get("busca_valor")
        tipo_busca = response.meta.get("tipo_busca")
        ean_entrada = response.meta.get("ean_entrada")
        link = response.meta.get("link_produto")
        sel = self._get_selector(response)

        nome = sel.css("h1::text").get()
        if nome:
            nome = " ".join(nome.split())
        else:
            nome = sel.css(
                ".vtex-store-components-3-x-productNameContainer *::text, "
                ".vtex-product-name-1-x-productName::text"
            ).get()
            if nome:
                nome = " ".join(nome.split())

        precos = self.extrair_precos_pdp(sel)
        status = "encontrado_com_preco" if precos.get("preco") else "encontrado_sem_preco"

        yield {
            "loja": "dia_ar",
            "tipo_busca": tipo_busca,
            "busca_valor": busca_valor,
            "ean": ean_entrada if tipo_busca in ("ean", "nome") else None,
            "nome": nome,
            "marca": self.extrair_marca(nome),
            "preco": precos.get("preco"),
            "precoPor": precos.get("precoPor"),
            "precoDe": precos.get("precoDe"),
            "oferta": precos.get("oferta"),
            "link": link or response.url,
            "status_busca": status,
        }