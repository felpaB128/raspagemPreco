import csv
import json
import re
import unicodedata
from pathlib import Path
from urllib.parse import quote, urljoin

import scrapy
from scrapy import Request

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from scrapy_playwright.page import PageMethod
except ImportError:
    PageMethod = None


class CarrefourPrecoSpider(scrapy.Spider):
    name = "carrefour_preco"
    allowed_domains = ["www.carrefour.com.ar", "carrefour.com.ar"]

    custom_settings = {
        "CONCURRENT_REQUESTS_PER_DOMAIN": 2,
        "DOWNLOAD_DELAY": 0.4,
        "DOWNLOAD_TIMEOUT": 60,
        "LOG_LEVEL": "INFO",
        "COOKIES_ENABLED": True,
    }

    PASTA_LEITURA = Path(r"C:\Users\felpa\OneDrive\Área de Trabalho\trabalho\leitura")
    ARQUIVO_PADRAO = PASTA_LEITURA / "BASE RETAIL MINORISTA 23_02_2026.xlsx"

    PROMO_PATTERNS = [
        {
            "tipo": "nxm",
            "regex": re.compile(
                r"\b(\d+)\s*[xX]\s*(\d+)\b|"
                r"\bleve\s+(\d+)\s+(?:e\s+)?pague\s+(\d+)\b|"
                r"\blleva\s+(\d+)\s+y?\s*paga\s+(\d+)\b|"
                r"\bna\s+compra\s+de\s+(\d+)\s+(?:unidades?\s+)?pague\s+(\d+)\b|"
                r"\bcomprando\s+(\d+)\s+(?:unidades?\s+)?pague\s+(\d+)\b",
                re.I,
            ),
        },
        {
            "tipo": "compre_leve",
            "regex": re.compile(
                r"\bcompre\s+(\d+)\s+(?:e\s+)?leve\s+(\d+)\b|"
                r"\bcompra\s+(\d+)\s+y?\s*lleva\s+(\d+)\b|"
                r"\blevando\s+(\d+)\s+(?:ganha|leve)\s+(\d+)\b|"
                r"\bll(evando|evando)\s+(\d+)\s+(?:te\s+)?llevas\s+(\d+)\b",
                re.I,
            ),
        },
        {
            "tipo": "ultimo_gratis",
            "regex": re.compile(
                r"\bna\s+compra\s+de\s+(\d+)\s+.*?(ultimo|último)\s+(?:sai\s+)?gratis\b|"
                r"\bcomprando\s+(\d+)\s+.*?(ultimo|último)\s+(?:sai\s+)?gratis\b|"
                r"\ben\s+la\s+compra\s+de\s+(\d+)\s+.*?(ultimo|último)\s+(?:sale\s+)?gratis\b|"
                r"\bllevando\s+(\d+)\s+.*?(ultimo|último)\s+(?:sale\s+)?gratis\b",
                re.I,
            ),
        },
        {
            "tipo": "segunda_unidade_percentual",
            "regex": re.compile(
                r"\b(2da|2do|segunda|segundo)\s+unidad(?:e)?\s*(?:a|al|com|con|c\/)?\s*(\d{1,3})\s*%\b|"
                r"\b(2da|2do|segunda|segundo)\s+unidad(?:e)?\s+(?:con|com)\s+(\d{1,3})\s*%\s+de\s+descuento\b|"
                r"\b(segunda|segundo)\s+unidade\s+com\s+(\d{1,3})\s*%\b",
                re.I,
            ),
        },
        {
            "tipo": "desconto_percentual",
            "regex": re.compile(
                r"\b(\d{1,3})\s*%\s*(?:off|dto|descuento|desconto)\b|"
                r"\bahorra\s+(\d{1,3})\s*%\b|"
                r"\beconomiza\s+(\d{1,3})\s*%\b",
                re.I,
            ),
        },
        {
            "tipo": "quantidade_minima",
            "regex": re.compile(
                r"\ba\s+partir\s+de\s+(\d+)\s+unidades?\b|"
                r"\bllevando\s+(\d+)\s+unidades?\b|"
                r"\bna\s+compra\s+de\s+(\d+)\s+unidades?\b|"
                r"\bcomprando\s+(\d+)\s+unidades?\b|"
                r"\b(\d+)\s+ou\s+mais\b|"
                r"\b(\d+)\s+o\s+mas\b",
                re.I,
            ),
        },
        {
            "tipo": "pack_combo",
            "regex": re.compile(
                r"\bpack\b|\bcombo\b|\bcombinable(?:s)?\b|\bcombinado(?:s)?\b|"
                r"\bahorro\b|\bpromo(?:cion)?\b",
                re.I,
            ),
        },
    ]

    PRICE_KEYS = {
        "price", "pricevalue", "pricenow", "sellingprice", "spotprice", "bestprice",
        "value", "saleprice", "pricetoview", "offerprice"
    }
    LIST_PRICE_KEYS = {
        "listprice", "pricewithoutdiscount", "referenceprice", "compareatprice",
        "strikethroughprice", "oldprice", "regularprice"
    }

    IGNORAR_TRECHOS_PRECO = [
        "cuotas sin interés",
        "cuotas sin interes",
        "csi",
        "mi carrefour crédito",
        "mi carrefour credito",
        "adicional",
        "off en 1 pago",
        "sin impuestos nacionales",
        "precio sin impuestos nacionales",
        "hot sale",
        "imperdibles",
        "envío gratis",
        "envio gratis",
    ]

    def __init__(self, arquivo_entrada=None, ean=None, usar_playwright="0", *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.ean = ean
        self.usar_playwright = str(usar_playwright).strip().lower() in ("1", "true", "yes", "sim")
        self.arquivo_entrada = Path(arquivo_entrada) if arquivo_entrada else self.ARQUIVO_PADRAO

    # ---------------- start ----------------

    def start_requests(self):
        eans = []

        if self.ean:
            ean = self.sanitize_ean(self.ean)
            if self.ean_valido(ean):
                eans.append(ean)
        elif self.arquivo_entrada:
            eans.extend(self.ler_eans_arquivo(self.arquivo_entrada))

        eans = list(dict.fromkeys([e for e in eans if self.ean_valido(e)]))

        for ean in eans:
            api_url = self.build_api_ean_url(ean)
            yield Request(
                api_url,
                callback=self.parse_api,
                meta={"ean": ean, "api_kind": "ean"},
                dont_filter=True,
            )

    # ---------------- leitura ----------------

    def resolver_caminho_arquivo(self, arquivo_entrada):
        caminho = Path(arquivo_entrada)

        if caminho.exists():
            return caminho

        if not caminho.is_absolute():
            candidato = self.PASTA_LEITURA / caminho.name
            if candidato.exists():
                return candidato

        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    def ler_eans_csv(self, caminho: Path):
        eans = []

        with caminho.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)

            if not reader.fieldnames:
                raise ValueError("CSV sem cabeçalho.")

            nomes = {c.lower().strip(): c for c in reader.fieldnames}
            coluna_ean = (
                nomes.get("ean")
                or nomes.get("código ean")
                or nomes.get("codigo ean")
                or nomes.get("codigo_ean")
                or nomes.get("codigoean")
                or nomes.get("ean 13")
                or nomes.get("cod ean")
                or nomes.get("cod_ean")
            )

            if not coluna_ean:
                raise ValueError(f"Não encontrei coluna EAN no CSV. Cabeçalho: {reader.fieldnames}")

            for row in reader:
                valor = (row.get(coluna_ean) or "").strip()
                if valor:
                    eans.append(self.sanitize_ean(valor))

        return list(dict.fromkeys(eans))

    def ler_eans_xlsx(self, caminho: Path):
        if openpyxl is None:
            raise RuntimeError("openpyxl não instalado. Rode: pip install openpyxl")

        wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Planilha vazia.")

        header_row_idx = None
        for i, row in enumerate(rows):
            if not row:
                continue
            vals = [str(c).strip().lower() if c is not None else "" for c in row]
            if any("ean" in v for v in vals):
                header_row_idx = i
                break

        if header_row_idx is None:
            raise ValueError("Não encontrei cabeçalho com EAN.")

        header = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
        header_norm = [h.lower() for h in header]
        nomes = {c: idx for idx, c in enumerate(header_norm)}

        idx_ean = None
        for chave in [
            "ean", "código ean", "codigo ean", "codigo_ean",
            "codigoean", "ean 13", "cod ean", "cod_ean"
        ]:
            if chave in nomes:
                idx_ean = nomes[chave]
                break

        if idx_ean is None:
            for nome_coluna, idx in nomes.items():
                if "ean" in nome_coluna:
                    idx_ean = idx
                    break

        if idx_ean is None:
            raise ValueError(f"Não encontrei coluna EAN no XLSX. Cabeçalho: {header}")

        eans = []
        for row in rows[header_row_idx + 1:]:
            if row is None or idx_ean >= len(row):
                continue
            valor = row[idx_ean]
            if valor is None:
                continue
            valor = self.sanitize_ean(valor)
            if valor:
                eans.append(valor)

        return list(dict.fromkeys(eans))

    def ler_eans_arquivo(self, arquivo_entrada):
        caminho = self.resolver_caminho_arquivo(arquivo_entrada)
        self.logger.info("Arquivo de entrada localizado em: %s", caminho)

        if caminho.suffix.lower() == ".csv":
            return self.ler_eans_csv(caminho)
        if caminho.suffix.lower() == ".xlsx":
            return self.ler_eans_xlsx(caminho)

        raise ValueError("Use arquivo .csv ou .xlsx")

    # ---------------- utils ----------------

    def sanitize_ean(self, value):
        return re.sub(r"\D", "", str(value or ""))

    def ean_valido(self, ean):
        return bool(ean) and len(ean) in (8, 12, 13, 14)

    def normalize_text(self, value):
        if value is None:
            return None
        value = re.sub(r"\s+", " ", str(value)).strip()
        return value or None

    def strip_accents(self, text):
        if not text:
            return ""
        return "".join(
            ch for ch in unicodedata.normalize("NFKD", str(text))
            if not unicodedata.combining(ch)
        )

    def normalize_promo_text(self, text):
        text = self.strip_accents(text or "")
        text = text.lower().replace("\xa0", " ")
        text = re.sub(r"\s+", " ", text).strip()
        return text

    def parse_price(self, value):
        if value is None:
            return None

        if isinstance(value, (int, float)):
            v = float(value)
            if v <= 0:
                return None
            return v

        texto = str(value)
        texto = texto.replace("\xa0", " ")
        texto = re.sub(r"\s+", "", texto)
        texto = texto.replace("$", "")

        m = re.search(r"(\d{1,3}(?:\.\d{3})+,\d{2}|\d{1,3}(?:\.\d{3})+|\d+,\d{2}|\d+\.\d{2}|\d+)", texto)
        if not m:
            return None

        texto = m.group(1)

        if "," in texto and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        else:
            if texto.count(".") > 1:
                texto = texto.replace(".", "")
            elif re.match(r"^\d{1,3}(?:\.\d{3})+$", texto):
                texto = texto.replace(".", "")

        try:
            v = float(texto)
            if v <= 0:
                return None
            return v
        except Exception:
            return None

    def build_api_ean_url(self, ean):
        fq = quote(f"alternateIds_Ean:{ean}", safe="")
        return f"https://www.carrefour.com.ar/api/catalog_system/pub/products/search?fq={fq}"

    def build_api_ft_url(self, ean):
        return f"https://www.carrefour.com.ar/api/catalog_system/pub/products/search?ft={quote(ean, safe='')}"

    def build_search_url(self, ean):
        return f"https://www.carrefour.com.ar/{quote(ean)}?_q={quote(ean)}&map=ft"

    def is_pdp_url(self, url):
        if not url:
            return False
        base = url.split("?")[0]
        return base.endswith("/p") or "/p/" in base or "idsku=" in url

    def abs_url(self, url):
        if not url:
            return None
        return urljoin("https://www.carrefour.com.ar", url)

    def get_default_seller(self, product):
        sellers = product.get("items", [{}])[0].get("sellers", [])
        if not sellers:
            return {}
        return next((s for s in sellers if s.get("sellerDefault")), sellers[0])

    def get_offer(self, product):
        seller = self.get_default_seller(product)
        return seller.get("commertialOffer") or {}

    def extract_json_ld_blocks(self, response):
        blocos = []
        for raw in response.css('script[type="application/ld+json"]::text').getall():
            raw = raw.strip()
            if not raw:
                continue
            try:
                data = json.loads(raw)
                blocos.append(data)
            except Exception:
                continue
        return blocos

    def find_offer_in_jsonld(self, data):
        if isinstance(data, dict):
            if data.get("@type") == "Offer":
                return data
            for v in data.values():
                achou = self.find_offer_in_jsonld(v)
                if achou:
                    return achou
        elif isinstance(data, list):
            for item in data:
                achou = self.find_offer_in_jsonld(item)
                if achou:
                    return achou
        return None

    def extract_embedded_json_objects(self, response):
        encontrados = []

        scripts = response.css("script::text").getall()
        for raw in scripts:
            if not raw:
                continue

            raw = raw.strip()
            if not raw:
                continue

            candidatos = [
                r'__NEXT_DATA__\s*=\s*({.*?})\s*;?\s*$',
                r'window\.__STATE__\s*=\s*({.*?})\s*;?\s*$',
                r'window\.__INITIAL_STATE__\s*=\s*({.*?})\s*;?\s*$',
                r'__PRELOADED_STATE__\s*=\s*({.*?})\s*;?\s*$',
            ]

            for pat in candidatos:
                m = re.search(pat, raw, re.S | re.M)
                if not m:
                    continue
                try:
                    encontrados.append(json.loads(m.group(1)))
                except Exception:
                    pass

            if raw.startswith("{") and raw.endswith("}"):
                try:
                    encontrados.append(json.loads(raw))
                except Exception:
                    pass

        return encontrados

    def walk_prices(self, obj, path="root", out=None):
        if out is None:
            out = []

        if isinstance(obj, dict):
            lowered = {str(k).lower(): v for k, v in obj.items()}

            for k, v in lowered.items():
                preco = self.parse_price(v)
                if preco is None:
                    continue

                if k in self.PRICE_KEYS:
                    out.append({"tipo": "precoPor", "valor": preco, "path": f"{path}.{k}"})
                elif k in self.LIST_PRICE_KEYS:
                    out.append({"tipo": "precoDe", "valor": preco, "path": f"{path}.{k}"})

            for k, v in obj.items():
                self.walk_prices(v, f"{path}.{k}", out)

        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                self.walk_prices(item, f"{path}[{i}]", out)

        return out

    def filtrar_texto_ruim_preco(self, texto):
        t = self.normalize_promo_text(texto)
        return any(x in t for x in self.IGNORAR_TRECHOS_PRECO)

    def is_unavailable_text(self, text):
        if not text:
            return False

        txt = self.normalize_promo_text(text)
        gatilhos = [
            "no disponible",
            "produto indisponivel",
            "producto no disponible",
            "sin stock",
            "agotado",
            "out of stock",
        ]
        return any(g in txt for g in gatilhos)

    def mark_unavailable(self, item):
        item["preco"] = 0.0
        item["precoPor"] = 0.0
        item["precoDe"] = 0.0
        item["oferta"] = None
        item["desconto_percentual"] = None
        item["indisponivel"] = 1
        return item

    def extract_price_candidates_from_text(self, text, skip_combo_promos=False):
        if not text:
            return []

        text = text.replace("\xa0", " ")
        text = re.sub(r"\s+", " ", text)

        padrao = r"\$\s*\d[\d\s\.\,]*\d"
        candidatos = []

        for match in re.finditer(padrao, text):
            trecho = match.group(0)
            janela = text[max(0, match.start()-80):min(len(text), match.end()+80)]

            if self.filtrar_texto_ruim_preco(janela):
                continue

            preco = self.parse_price(trecho)
            if preco is not None:
                candidatos.append(preco)

        unicos = []
        for v in candidatos:
            if v not in unicos:
                unicos.append(v)
        return unicos

    def detect_promotions(self, text):
        if not text:
            return []

        texto = self.normalize_promo_text(text)
        encontrados = []

        for spec in self.PROMO_PATTERNS:
            for m in spec["regex"].finditer(texto):
                trecho = texto[max(0, m.start() - 100):min(len(texto), m.end() + 140)]
                precos = self.extract_price_candidates_from_text(trecho, skip_combo_promos=False)
                grupos = [g for g in m.groups() if g is not None]

                promo = {
                    "tipo": spec["tipo"],
                    "match": m.group(0),
                    "texto_contexto": trecho[:350],
                    "grupos": grupos,
                    "preco_unitario": precos[0] if precos else None,
                    "preco_total": precos[1] if len(precos) > 1 else None,
                }

                nums = [int(g) for g in grupos if str(g).isdigit()]

                if spec["tipo"] == "nxm" and len(nums) >= 2:
                    promo["qtd_leva"] = nums[0]
                    promo["qtd_paga"] = nums[1]
                    promo["mecanica"] = f"leve_{nums[0]}_pague_{nums[1]}"

                elif spec["tipo"] == "compre_leve" and len(nums) >= 2:
                    promo["qtd_compra"] = nums[0]
                    promo["qtd_leva"] = nums[1]
                    promo["mecanica"] = f"compre_{nums[0]}_leve_{nums[1]}"

                elif spec["tipo"] == "ultimo_gratis" and len(nums) >= 1:
                    promo["qtd_acionadora"] = nums[0]
                    promo["mecanica"] = f"na_compra_de_{nums[0]}_ultimo_gratis"

                elif spec["tipo"] == "segunda_unidade_percentual" and len(nums) >= 1:
                    promo["percentual"] = nums[-1]
                    promo["mecanica"] = f"segunda_unidade_{nums[-1]}_percent"

                elif spec["tipo"] == "desconto_percentual" and len(nums) >= 1:
                    promo["percentual"] = nums[0]
                    promo["mecanica"] = f"desconto_{nums[0]}_percent"

                elif spec["tipo"] == "quantidade_minima" and len(nums) >= 1:
                    promo["quantidade_minima"] = nums[0]
                    promo["mecanica"] = f"quantidade_minima_{nums[0]}"

                elif spec["tipo"] == "pack_combo":
                    promo["mecanica"] = "pack_combo"

                encontrados.append(promo)

        unicos = []
        vistos = set()
        for p in encontrados:
            chave = (p.get("tipo"), p.get("match"))
            if chave not in vistos:
                vistos.add(chave)
                unicos.append(p)

        return unicos

    def extract_prices_from_dom(self, response):
        blocos = []

        seletores_precisos = [
            "[class*='sellingPrice'] *::text",
            "[class*='listPrice'] *::text",
            "[class*='priceContainer'] *::text",
            "[class*='price_'] *::text",
            "[data-testid*='price'] *::text",
            "[class*='spotPrice'] *::text",
            "[class*='productPrice'] *::text",
            "[class*='currencyContainer'] *::text",
            "[class*='savings'] *::text",
            "[class*='promo'] *::text",
            "[class*='discount'] *::text",
            "[class*='benefit'] *::text",
            "section *::text",
            "article *::text",
        ]

        for sel in seletores_precisos:
            vals = response.css(sel).getall()
            if vals:
                texto = " ".join([v.strip() for v in vals if v and v.strip()])
                texto = re.sub(r"\s+", " ", texto).strip()
                if texto:
                    blocos.append(texto)

        blocos_filtrados = []
        for b in blocos:
            if len(b) < 4:
                continue
            blocos_filtrados.append(b)

        melhores_precos = []
        todas_promos = []

        for bloco in blocos_filtrados[:80]:
            precos = self.extract_price_candidates_from_text(bloco)
            promos = self.detect_promotions(bloco)

            if precos:
                melhores_precos.append({
                    "texto": bloco[:500],
                    "precos": precos[:5],
                    "qtd": len(precos),
                })

            if promos:
                todas_promos.extend(promos)

        preco_por = None
        preco_de = None

        if melhores_precos:
            candidatos = []
            for reg in melhores_precos:
                for p in reg["precos"]:
                    candidatos.append(p)

            candidatos = [x for x in candidatos if x is not None]
            candidatos_unicos = []
            for x in candidatos:
                if x not in candidatos_unicos:
                    candidatos_unicos.append(x)

            if candidatos_unicos:
                preco_por = min(candidatos_unicos)
                maior = max(candidatos_unicos)
                if maior > preco_por:
                    preco_de = maior

        promos_unicas = []
        vistos = set()
        for p in todas_promos:
            chave = (p.get("tipo"), p.get("match"))
            if chave not in vistos:
                vistos.add(chave)
                promos_unicas.append(p)

        texto_debug = " || ".join([x["texto"] for x in melhores_precos[:8]])[:2000] if melhores_precos else None

        return {
            "texto": texto_debug,
            "precoPor": preco_por,
            "precoDe": preco_de,
            "combo": promos_unicas[0] if promos_unicas else None,
            "promocoes": promos_unicas,
        }

    def pick_product_from_api(self, products, ean):
        if not isinstance(products, list):
            return None

        for product in products:
            for item in product.get("items", []) or []:
                eans = [
                    self.sanitize_ean(item.get("ean")),
                    self.sanitize_ean(product.get("productReference")),
                    self.sanitize_ean(product.get("productReferenceCode")),
                ]
                if ean in eans:
                    return product

        return products[0] if products else None

    def extract_product_fields_from_api(self, product):
        if not product:
            return {}

        item0 = (product.get("items") or [{}])[0]
        offer = self.get_offer(product)

        link = (
            product.get("link")
            or product.get("linkText")
            or product.get("detailUrl")
        )
        if link and not str(link).startswith("http"):
            if not str(link).startswith("/"):
                link = "/" + str(link)
            link = self.abs_url(link)

        out = {
            "nome": self.normalize_text(product.get("productName") or product.get("name")),
            "marca": self.normalize_text(product.get("brand")),
            "sku": item0.get("itemId"),
            "link": link,
            "price_raw": offer.get("Price") or offer.get("spotPrice"),
            "list_price_raw": offer.get("ListPrice"),
            "price_without_discount_raw": offer.get("PriceWithoutDiscount") or offer.get("priceWithoutDiscount"),
        }
        return out

    def normalize_prices(self, item):
        preco = self.parse_price(item.get("preco") or item.get("precoPor"))
        preco_por = self.parse_price(item.get("precoPor") or item.get("preco"))
        preco_de = self.parse_price(item.get("precoDe"))

        if item.get("indisponivel") == 1:
            item["preco"] = 0.0
            item["precoPor"] = 0.0
            item["precoDe"] = 0.0
            item["oferta"] = None
            item["desconto_percentual"] = None
            return item

        promo_tipo = item.get("promo_tipo")
        promo_texto = item.get("promo_texto")
        promo_preco_unitario = self.parse_price(item.get("promo_preco_unitario"))
        promo_preco_total = self.parse_price(item.get("promo_preco_total"))

        if preco is None and preco_por is not None:
            preco = preco_por
        if preco_por is None and preco is not None:
            preco_por = preco

        if preco_de is not None and preco_por is not None and preco_de <= preco_por:
            preco_de = None

        oferta = None
        desconto_percentual = None

        if preco_de is not None and preco_por is not None and preco_de > preco_por:
            oferta = "x"
            try:
                desconto_percentual = round((1 - (preco_por / preco_de)) * 100, 2)
            except Exception:
                desconto_percentual = None

        item["preco"] = preco
        item["precoPor"] = preco_por
        item["precoDe"] = preco_de
        item["oferta"] = oferta
        item["desconto_percentual"] = desconto_percentual

        item["promo_tipo"] = promo_tipo
        item["promo_texto"] = promo_texto
        item["promo_preco_unitario"] = promo_preco_unitario
        item["promo_preco_total"] = promo_preco_total

        return item

    # ---------------- parse api ----------------

    def parse_api(self, response):
        ean = response.meta["ean"]

        try:
            data = json.loads(response.text)
        except Exception:
            data = None

        product = self.pick_product_from_api(data, ean) if data else None

        if not product:
            ft_url = self.build_api_ft_url(ean)
            yield Request(
                ft_url,
                callback=self.parse_api_ft,
                meta={"ean": ean},
                dont_filter=True,
            )
            return

        base = self.extract_product_fields_from_api(product)

        item = {
            "loja": "Carrefour",
            "ean": ean,
            "sku": base.get("sku"),
            "nome": base.get("nome"),
            "marca": base.get("marca"),
            "preco": base.get("price_raw"),
            "precoPor": base.get("price_raw"),
            "precoDe": base.get("list_price_raw") or base.get("price_without_discount_raw"),
            "oferta": None,
            "desconto_percentual": None,
            "promo_tipo": None,
            "promo_texto": None,
            "promo_preco_unitario": None,
            "promo_preco_total": None,
            "promo_mecanica": None,
            "promo_percentual": None,
            "promo_qtd_leva": None,
            "promo_qtd_paga": None,
            "promo_qtd_compra": None,
            "promo_quantidade_minima": None,
            "promocoes_detectadas": None,
            "price_raw": base.get("price_raw"),
            "list_price_raw": base.get("list_price_raw"),
            "price_without_discount_raw": base.get("price_without_discount_raw"),
            "link": base.get("link"),
            "indisponivel": 0,
        }

        item = self.normalize_prices(item)

        pdp_url = item.get("link") or self.build_search_url(ean)

        if self.usar_playwright and PageMethod is not None:
            yield Request(
                pdp_url,
                callback=self.parse_pdp_playwright,
                meta={
                    "ean": ean,
                    "item_base": item,
                    "playwright": True,
                    "playwright_include_page": True,
                    "playwright_page_methods": [
                        PageMethod("wait_for_load_state", "networkidle"),
                    ],
                },
                errback=self.errback_close_page,
                dont_filter=True,
            )
        else:
            yield Request(
                pdp_url,
                callback=self.parse_pdp,
                meta={"ean": ean, "item_base": item},
                dont_filter=True,
            )

    def parse_api_ft(self, response):
        ean = response.meta["ean"]

        try:
            data = json.loads(response.text)
        except Exception:
            data = None

        product = self.pick_product_from_api(data, ean) if data else None

        if not product:
            yield {
                "loja": "Carrefour",
                "ean": ean,
                "sku": None,
                "nome": None,
                "marca": None,
                "preco": None,
                "precoPor": None,
                "precoDe": None,
                "oferta": None,
                "desconto_percentual": None,
                "promo_tipo": None,
                "promo_texto": None,
                "promo_preco_unitario": None,
                "promo_preco_total": None,
                "promo_mecanica": None,
                "promo_percentual": None,
                "promo_qtd_leva": None,
                "promo_qtd_paga": None,
                "promo_qtd_compra": None,
                "promo_quantidade_minima": None,
                "promocoes_detectadas": None,
                "price_raw": None,
                "list_price_raw": None,
                "price_without_discount_raw": None,
                "link": self.build_search_url(ean),
                "indisponivel": 0,
            }
            return

        base = self.extract_product_fields_from_api(product)

        item = {
            "loja": "Carrefour",
            "ean": ean,
            "sku": base.get("sku"),
            "nome": base.get("nome"),
            "marca": base.get("marca"),
            "preco": base.get("price_raw"),
            "precoPor": base.get("price_raw"),
            "precoDe": base.get("list_price_raw") or base.get("price_without_discount_raw"),
            "oferta": None,
            "desconto_percentual": None,
            "promo_tipo": None,
            "promo_texto": None,
            "promo_preco_unitario": None,
            "promo_preco_total": None,
            "promo_mecanica": None,
            "promo_percentual": None,
            "promo_qtd_leva": None,
            "promo_qtd_paga": None,
            "promo_qtd_compra": None,
            "promo_quantidade_minima": None,
            "promocoes_detectadas": None,
            "price_raw": base.get("price_raw"),
            "list_price_raw": base.get("list_price_raw"),
            "price_without_discount_raw": base.get("price_without_discount_raw"),
            "link": base.get("link"),
            "indisponivel": 0,
        }

        item = self.normalize_prices(item)

        pdp_url = item.get("link") or self.build_search_url(ean)

        yield Request(
            pdp_url,
            callback=self.parse_pdp,
            meta={"ean": ean, "item_base": item},
            dont_filter=True,
        )

    # ---------------- parse pdp ----------------

    def merge_structured_price_sources(self, response, item):
        candidatos = []

        for bloco in self.extract_json_ld_blocks(response):
            candidatos.extend(self.walk_prices(bloco, path="jsonld"))

            offer = self.find_offer_in_jsonld(bloco)
            if offer:
                p = self.parse_price(offer.get("price"))
                if p is not None:
                    candidatos.append({"tipo": "precoPor", "valor": p, "path": "jsonld.offer.price"})

                ps = offer.get("priceSpecification")
                if isinstance(ps, dict):
                    lp = self.parse_price(ps.get("price"))
                    price_type = str(ps.get("priceType") or "")
                    if lp is not None and "StrikethroughPrice" in price_type:
                        candidatos.append({"tipo": "precoDe", "valor": lp, "path": "jsonld.offer.priceSpecification"})
                elif isinstance(ps, list):
                    for i, spec in enumerate(ps):
                        if not isinstance(spec, dict):
                            continue
                        lp = self.parse_price(spec.get("price"))
                        price_type = str(spec.get("priceType") or "")
                        if lp is not None and "StrikethroughPrice" in price_type:
                            candidatos.append({"tipo": "precoDe", "valor": lp, "path": f"jsonld.offer.priceSpecification[{i}]"})

        for emb in self.extract_embedded_json_objects(response):
            candidatos.extend(self.walk_prices(emb, path="embedded"))

        preco_por = item.get("precoPor")
        preco_de = item.get("precoDe")

        for c in candidatos:
            if c["tipo"] == "precoPor" and preco_por is None:
                preco_por = c["valor"]

        lista_preco_de = [c["valor"] for c in candidatos if c["tipo"] == "precoDe"]
        lista_preco_por = [c["valor"] for c in candidatos if c["tipo"] == "precoPor"]

        if preco_por is None and lista_preco_por:
            preco_por = min(lista_preco_por)

        if lista_preco_de:
            cand = max(lista_preco_de)
            if preco_por is None or cand > preco_por:
                preco_de = cand

        item["preco"] = preco_por if preco_por is not None else item.get("preco")
        item["precoPor"] = preco_por
        item["precoDe"] = preco_de
        item["debug_candidatos_estruturados"] = candidatos[:50] or None
        return item

    def merge_pdp_data(self, response, item):
        texto_pagina = " ".join(
            response.css("h1 *::text, section *::text, article *::text, [class*='product'] *::text").getall()
        )
        texto_pagina = re.sub(r"\s+", " ", texto_pagina).strip()

        if self.is_unavailable_text(texto_pagina):
            item["link"] = response.url
            item["debug_texto_preco"] = texto_pagina[:700] if texto_pagina else None
            return self.mark_unavailable(item)

        item = self.merge_structured_price_sources(response, item)

        dom = self.extract_prices_from_dom(response)
        texto_dom = dom.get("texto")
        promocoes = dom.get("promocoes") or []

        dom_preco_por = self.parse_price(dom.get("precoPor"))
        dom_preco_de = self.parse_price(dom.get("precoDe"))

        if item.get("precoPor") is None and dom_preco_por is not None:
            item["preco"] = dom_preco_por
            item["precoPor"] = dom_preco_por

        if item.get("precoDe") is None and dom_preco_de is not None:
            item["precoDe"] = dom_preco_de

        if promocoes:
            item["promocoes_detectadas"] = promocoes
            principal = promocoes[0]

            item["promo_tipo"] = principal.get("tipo")
            item["promo_texto"] = principal.get("texto_contexto") or principal.get("match")
            item["promo_preco_unitario"] = principal.get("preco_unitario")
            item["promo_preco_total"] = principal.get("preco_total")
            item["promo_mecanica"] = principal.get("mecanica")
            item["promo_percentual"] = principal.get("percentual")
            item["promo_qtd_leva"] = principal.get("qtd_leva")
            item["promo_qtd_paga"] = principal.get("qtd_paga")
            item["promo_qtd_compra"] = principal.get("qtd_compra")
            item["promo_quantidade_minima"] = principal.get("quantidade_minima")

        if not item.get("nome"):
            nome = response.css("h1 *::text").getall()
            if nome:
                item["nome"] = self.normalize_text(" ".join(nome))

        item["link"] = response.url
        item["debug_texto_preco"] = texto_dom[:700] if texto_dom else None

        return self.normalize_prices(item)

    def parse_pdp(self, response):
        item = dict(response.meta["item_base"])
        item = self.merge_pdp_data(response, item)
        yield item

    async def parse_pdp_playwright(self, response):
        item = dict(response.meta["item_base"])
        page = response.meta.get("playwright_page")

        try:
            item = self.merge_pdp_data(response, item)
            yield item
        finally:
            if page:
                await page.close()

    async def errback_close_page(self, failure):
        page = failure.request.meta.get("playwright_page")
        if page:
            await page.close()