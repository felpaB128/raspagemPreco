import csv
import re
from pathlib import Path
from urllib.parse import quote

from scrapy import Spider, Request, Selector

try:
    import openpyxl
except ImportError:
    openpyxl = None


class MasonlinePrecoSpider(Spider):
    name = "masonline_preco"
    allowed_domains = ["www.masonline.com.ar", "masonline.com.ar"]

    custom_settings = {
        "ZYTE_API_TRANSPARENT_MODE": True,
        "CONCURRENT_REQUESTS_PER_DOMAIN": 4,
        "DOWNLOAD_DELAY": 0.2,
        "LOG_LEVEL": "INFO",
        "FEED_EXPORT_ENCODING": "utf-8",
    }

    def __init__(self, ean=None, arquivo_entrada=None, termo=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if not ean and not arquivo_entrada and not termo:
            raise ValueError(
                "Passe ean, termo ou arquivo_entrada. "
                "Ex.: -a ean=190679000019 | -a termo=celula | -a arquivo_entrada=eans.csv"
            )

        self.ean = ean
        self.termo = termo
        self.arquivo_entrada = arquivo_entrada

        # chave: (tipo_busca, busca_valor) -> lista de itens brutos
        self.resultados = {}

    # -----------------------------
    # leitura de arquivo
    # -----------------------------
    def _ler_eans_csv(self, caminho: Path):
        eans = []
        with caminho.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("CSV sem cabeçalho.")

            nomes = {c.lower().strip(): c for c in reader.fieldnames}
            coluna_ean = (
                nomes.get("ean")
                or nomes.get("código ean")
                or nomes.get("codigo ean")
                or nomes.get("codigo_ean")
                or nomes.get("codigoean")
                or nomes.get("ean 13")
                or nomes.get("cod ean")
                or nomes.get("cod_ean")
            )

            if not coluna_ean:
                raise ValueError(
                    f"Não encontrei coluna de EAN no CSV. Cabeçalho: {reader.fieldnames}"
                )

            for row in reader:
                valor = (row.get(coluna_ean) or "").strip()
                if valor:
                    eans.append(valor)

        return list(dict.fromkeys(eans))

    def _ler_eans_xlsx(self, caminho: Path):
        if openpyxl is None:
            raise RuntimeError("openpyxl não está instalado. Rode: pip install openpyxl")

        wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Planilha vazia.")

        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        header_norm = [h.lower() for h in header]
        nomes = {c: idx for idx, c in enumerate(header_norm)}

        idx_ean = (
            nomes.get("ean")
            or nomes.get("código ean")
            or nomes.get("codigo ean")
            or nomes.get("codigo_ean")
            or nomes.get("codigoean")
            or nomes.get("ean 13")
            or nomes.get("cod ean")
            or nomes.get("cod_ean")
        )

        if idx_ean is None:
            raise ValueError(f"Não encontrei coluna de EAN no XLSX. Cabeçalho: {header}")

        eans = []
        for row in rows[1:]:
            if not row or idx_ean >= len(row):
                continue
            valor = row[idx_ean]
            if valor is None:
                continue
            valor = str(valor).strip()
            if valor:
                eans.append(valor)

        return list(dict.fromkeys(eans))

    def _resolver_caminho_arquivo(self, caminho_str: str) -> Path:
        caminho = Path(caminho_str).expanduser()

        if caminho.is_absolute():
            return caminho

        candidatos = [
            Path.cwd() / caminho,
            Path(__file__).resolve().parents[3] / caminho,
            Path(__file__).resolve().parent / caminho,
        ]

        for candidato in candidatos:
            if candidato.exists():
                return candidato

        return candidatos[0]

    def _ler_eans_arquivo(self, caminho_str: str):
        caminho = self._resolver_caminho_arquivo(caminho_str)

        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

        if caminho.suffix.lower() == ".csv":
            return self._ler_eans_csv(caminho)
        elif caminho.suffix.lower() == ".xlsx":
            return self._ler_eans_xlsx(caminho)

        raise ValueError("Arquivo de entrada deve ser .csv ou .xlsx")

    # -----------------------------
    # util
    # -----------------------------
    def montar_url_busca(self, termo: str) -> str:
        termo = str(termo).strip()
        termo_enc = quote(termo)
        return f"https://www.masonline.com.ar/{termo_enc}?map=ft&_q={termo_enc}"

    def _get_selector(self, response):
        raw = getattr(response, "raw_api_response", None) or {}
        browser_html = raw.get("browserHtml")
        if browser_html:
            return Selector(text=browser_html)
        return response

    def extrair_preco(self, texto: str):
        if not texto:
            return None

        texto = " ".join(texto.split())

        padroes = [
            r"\$\s*\d{1,3}(?:\.\d{3})*(?:,\d{2})?",
            r"R\$\s*\d{1,3}(?:\.\d{3})*,\d{2}",
        ]

        for padrao in padroes:
            m = re.search(padrao, texto, flags=re.IGNORECASE)
            if m:
                return m.group(0)

        return None

    def extrair_marca(self, nome):
        if not nome:
            return None
        partes = nome.split()
        return partes[0] if partes else None

    def extrair_nome_da_listagem(self, card):
        seletores = [
            "h2::text",
            "h3::text",
            "span::text",
            "a::text",
        ]
        for sel in seletores:
            txts = card.css(sel).getall()
            txts = [" ".join(t.split()) for t in txts if t and t.strip()]
            txts = [t for t in txts if len(t) > 2]
            if txts:
                return txts[0]
        return None

    # -----------------------------
    # helpers de consolidação
    # -----------------------------
    def _acumular_resultado(self, item):
        chave = (item["tipo_busca"], item["busca_valor"])
        self.resultados.setdefault(chave, []).append(item)

    def _limpar_string(self, s):
        if s is None:
            return None
        s = " ".join(str(s).split())
        return s or None

    def _consolidar_resultado(self, tipo_busca, busca_valor):
        itens = self.resultados.get((tipo_busca, busca_valor), []) or []

        if not itens:
            return None

        for it in itens:
            it["nome"] = self._limpar_string(it.get("nome"))
            it["marca"] = self._limpar_string(it.get("marca"))
            it["preco"] = self._limpar_string(it.get("preco"))
            it["link"] = self._limpar_string(it.get("link"))

        itens_ok_nome = [
            it
            for it in itens
            if it.get("nome")
            and "No encontramos ningún resultado" not in it["nome"]
        ]

        candidatos = itens_ok_nome or itens

        candidatos_preco_ok = [
            it
            for it in candidatos
            if it.get("preco") and it["preco"] != "$ 0"
        ]

        if candidatos_preco_ok:
            escolhido = candidatos_preco_ok[0]
        else:
            escolhido = candidatos[0]

        item_final = {
            "loja": "masonline_ar",
            "tipo_busca": tipo_busca,
            "busca_valor": busca_valor,
            "ean": busca_valor if tipo_busca == "ean" else None,
            "nome": escolhido.get("nome"),
            "marca": escolhido.get("marca"),
            "preco": escolhido.get("preco"),
            "link": escolhido.get("link"),
            "status_busca": escolhido.get("status_busca") or "consolidado",
        }
        return item_final

    # -----------------------------
    # start
    # -----------------------------
    async def start(self):
        if self.arquivo_entrada:
            caminho_resolvido = self._resolver_caminho_arquivo(self.arquivo_entrada)
            self.logger.info("Lendo arquivo de entrada: %s", caminho_resolvido)

            eans = self._ler_eans_arquivo(self.arquivo_entrada)[:50]
            self.logger.info("Processando %d EANs do arquivo", len(eans))

            for ean in eans:
                yield Request(
                    url=self.montar_url_busca(ean),
                    callback=self.parse_search,
                    dont_filter=True,
                    meta={
                        "busca_valor": str(ean),
                        "tipo_busca": "ean",
                        "zyte_api_automap": {
                            "browserHtml": True,
                            "actions": [{"action": "scrollBottom"}],
                        },
                    },
                )
            return

        valor_busca = self.termo or self.ean
        tipo_busca = "termo" if self.termo else "ean"

        yield Request(
            url=self.montar_url_busca(valor_busca),
            callback=self.parse_search,
            dont_filter=True,
            meta={
                "busca_valor": str(valor_busca),
                "tipo_busca": tipo_busca,
                "zyte_api_automap": {
                    "browserHtml": True,
                    "actions": [{"action": "scrollBottom"}],
                },
            },
        )

    # -----------------------------
    # parse search
    # -----------------------------
    def parse_search(self, response):
        busca_valor = response.meta.get("busca_valor")
        tipo_busca = response.meta.get("tipo_busca")
        sel = self._get_selector(response)

        self.logger.info(
            "Busca %s | valor=%s | URL=%s", tipo_busca, busca_valor, response.url
        )

        links = sel.css("a::attr(href)").getall()
        links_produto = []

        for href in links:
            if not href:
                continue
            href = href.strip()

            if (
                href.endswith("/p")
                or "/p?" in href
                or "/produto/" in href.lower()
                or "/product/" in href.lower()
            ):
                abs_url = response.urljoin(href)
                if abs_url not in links_produto:
                    links_produto.append(abs_url)

        self.logger.info("Links de produto encontrados: %d", len(links_produto))

        if links_produto:
            for link in links_produto:
                yield Request(
                    url=link,
                    callback=self.parse_produto,
                    dont_filter=True,
                    meta={
                        "busca_valor": busca_valor,
                        "tipo_busca": tipo_busca,
                        "link_produto": link,
                        "zyte_api_automap": {
                            "browserHtml": True,
                        },
                    },
                )
            return

        cards = sel.css("article, section, div")
        encontrados_listagem = 0

        for card in cards:
            html_card = card.get() or ""
            if "/p" not in html_card and "$" not in html_card:
                continue

            nome = self.extrair_nome_da_listagem(card)
            texto_card = " ".join(
                t.strip() for t in card.css("::text").getall() if t and t.strip()
            )
            preco = self.extrair_preco(texto_card)

            if nome or preco:
                encontrados_listagem += 1
                item = {
                    "loja": "masonline_ar",
                    "tipo_busca": tipo_busca,
                    "busca_valor": busca_valor,
                    "ean": busca_valor if tipo_busca == "ean" else None,
                    "nome": nome,
                    "marca": self.extrair_marca(nome),
                    "preco": preco,
                    "link": response.url,
                    "status_busca": "resultado_em_listagem",
                }
                self._acumular_resultado(item)

        if encontrados_listagem == 0:
            item = {
                "loja": "masonline_ar",
                "tipo_busca": tipo_busca,
                "busca_valor": busca_valor,
                "ean": busca_valor if tipo_busca == "ean" else None,
                "nome": "No encontramos ningún resultado",
                "marca": "No",
                "preco": None,
                "link": response.url,
                "status_busca": "nao_indexado_na_busca",
            }
            self._acumular_resultado(item)

        # terminou a busca desse valor → emite item consolidado
        consolidado = self._consolidar_resultado(tipo_busca, busca_valor)
        if consolidado:
            yield consolidado

    # -----------------------------
    # parse produto
    # -----------------------------
    def parse_produto(self, response):
        busca_valor = response.meta.get("busca_valor")
        tipo_busca = response.meta.get("tipo_busca")
        link = response.meta.get("link_produto")
        sel = self._get_selector(response)

        nome = sel.css("h1::text").get()
        if nome:
            nome = " ".join(nome.split())

        textos = [t.strip() for t in sel.css("body ::text").getall() if t and t.strip()]
        bloco = " ".join(textos)

        preco = self.extrair_preco(bloco)

        item = {
            "loja": "masonline_ar",
            "tipo_busca": tipo_busca,
            "busca_valor": busca_valor,
            "ean": busca_valor if tipo_busca == "ean" else None,
            "nome": nome,
            "marca": self.extrair_marca(nome),
            "preco": preco,
            "link": link,
            "status_busca": "encontrado",
        }
        self._acumular_resultado(item)
        # aqui não faz yield; quem consolidará e dará yield é parse_search