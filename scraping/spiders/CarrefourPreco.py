import csv
import json
import re
from pathlib import Path
from urllib.parse import quote

from scrapy import Request, Spider

try:
    import openpyxl
except ImportError:
    openpyxl = None


class CarrefourPrecoSpider(Spider):
    # nome ajustado para não conflitar com o spider do Makro
    name = "carrefour_preco"
    allowed_domains = ["www.carrefour.com.ar", "carrefour.com.ar"]

    custom_settings = {
        "CONCURRENT_REQUESTS_PER_DOMAIN": 4,
        "DOWNLOAD_DELAY": 0.2,
        "LOG_LEVEL": "INFO",
    }

    PALAVRAS_RUINS = [
        "off",
        "descuento",
        "descuentos",
        "promocion",
        "promoción",
        "promociones",
        "oferta",
        "ofertas",
        "seleccionados",
        "categorias",
        "categorías",
        "semana",
        "billetera",
        "beneficio",
        "ahorro",
        "hasta ",
    ]

    def __init__(self, arquivo_entrada=None, ean=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if not arquivo_entrada and not ean:
            raise ValueError("Passe arquivo_entrada ou ean")

        self.arquivo_entrada = arquivo_entrada
        self.ean = ean

    # ---------- LEITURA ARQUIVOS ----------

    def ler_eans_csv(self, caminho: Path):
        eans = []

        with caminho.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)

            if not reader.fieldnames:
                raise ValueError("CSV sem cabeçalho.")

            nomes = {c.lower().strip(): c for c in reader.fieldnames}
            coluna_ean = (
                nomes.get("ean")
                or nomes.get("código ean")
                or nomes.get("codigo ean")
                or nomes.get("codigo_ean")
                or nomes.get("codigoean")
                or nomes.get("ean 13")
                or nomes.get("cod ean")
                or nomes.get("cod_ean")
            )

            if not coluna_ean:
                raise ValueError(
                    f"Não encontrei coluna EAN no CSV. Cabeçalho: {reader.fieldnames}"
                )

            for row in reader:
                valor = (row.get(coluna_ean) or "").strip()
                if valor:
                    eans.append(valor)

        return list(dict.fromkeys(eans))

    def ler_eans_xlsx(self, caminho: Path):
        if openpyxl is None:
            raise RuntimeError("openpyxl não instalado. Rode: pip install openpyxl")

        wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Planilha vazia.")

        header_row_idx = None
        for i, row in enumerate(rows):
            if row and any(c is not None and str(c).strip() for c in row):
                header_row_idx = i
                break

        if header_row_idx is None:
            raise ValueError("Planilha sem dados.")

        header = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
        header_normalizado = [h.lower() for h in header]
        nomes = {c: idx for idx, c in enumerate(header_normalizado)}

        idx_ean = (
            nomes.get("ean")
            or nomes.get("código ean")
            or nomes.get("codigo ean")
            or nomes.get("codigo_ean")
            or nomes.get("codigoean")
            or nomes.get("ean 13")
            or nomes.get("cod ean")
            or nomes.get("cod_ean")
        )

        if idx_ean is None:
            for nome_coluna, idx in nomes.items():
                if "ean" in nome_coluna:
                    idx_ean = idx
                    break

        if idx_ean is None:
            raise ValueError(f"Não encontrei coluna EAN no XLSX. Cabeçalho: {header}")

        eans = []
        for row in rows[header_row_idx + 1:]:
            if row is None or idx_ean >= len(row):
                continue

            valor = row[idx_ean]
            if valor is None:
                continue

            valor = str(valor).strip()
            if valor:
                eans.append(valor)

        return list(dict.fromkeys(eans))

    def ler_eans_arquivo(self, arquivo_entrada):
        caminho = Path(arquivo_entrada)

        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

        if caminho.suffix.lower() == ".csv":
            eans = self.ler_eans_csv(caminho)
        elif caminho.suffix.lower() == ".xlsx":
            eans = self.ler_eans_xlsx(caminho)
        else:
            raise ValueError("Use arquivo .csv ou .xlsx")

        return eans[:50]

    # ---------- UTILS ----------

    def walk_dicts(self, obj):
        encontrados = []

        def walk(x):
            if isinstance(x, dict):
                encontrados.append(x)
                for v in x.values():
                    walk(v)
            elif isinstance(x, list):
                for i in x:
                    walk(i)

        walk(obj)
        return encontrados

    def parse_price(self, value):
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return float(value)

        texto = str(value).strip()
        texto = texto.replace("$", "").replace("\xa0", " ")
        texto = texto.replace(".", "").replace(",", ".")
        texto = re.sub(r"[^\d.]", "", texto)

        try:
            return float(texto) if texto else None
        except Exception:
            return None

    def normalize_text(self, value):
        if value is None:
            return None
        value = re.sub(r"\s+", " ", str(value)).strip()
        return value or None

    def build_search_url(self, ean):
        ean = str(ean).strip()
        return f"https://www.carrefour.com.ar/{quote(ean)}?_q={quote(ean)}&map=ft"

    def texto_ruim(self, texto):
        if not texto:
            return False
        t = texto.lower()
        return any(p in t for p in self.PALAVRAS_RUINS)

    def clean_nome(self, nome, ean):
        nome = self.normalize_text(nome)
        if not nome:
            return None

        nome_limpo = re.sub(r"\s*-\s*Carrefour\s*$", "", nome, flags=re.I).strip()

        if ean and re.sub(r"\D", "", nome_limpo) == str(ean):
            return None
        if ean and nome_limpo == f"{ean} - Carrefour":
            return None

        if self.texto_ruim(nome_limpo):
            return None

        return nome_limpo

    def item_score(self, item):
        nome = (item.get("nome") or "").strip()
        marca = (item.get("marca") or "").strip()
        link = (item.get("link") or "").strip().lower()
        preco = item.get("preco")

        score = 0

        if preco is not None:
            score += 100

        if nome:
            score += 20
            score += min(len(nome), 120)

        if marca:
            score += 10

        if link.endswith("/p") or "/p" in link:
            score += 120

        if "busca" in link or "_q=" in link or "map=ft" in link:
            score -= 40

        if self.texto_ruim(nome):
            score -= 200

        return score

    def add_candidate(self, candidatos, vistos, item):
        ean = self.normalize_text(item.get("ean"))
        nome = self.clean_nome(item.get("nome"), ean)

        normalizado = {
            "ean": ean,
            "nome": nome,
            "marca": self.normalize_text(item.get("marca")),
            "preco": self.parse_price(item.get("preco")),
            "loja": "carrefour_ar",
            "link": self.normalize_text(item.get("link")),
        }

        if not normalizado["nome"] and normalizado["preco"] is None:
            return

        chave_unica = json.dumps(normalizado, sort_keys=True, ensure_ascii=False)
        if chave_unica in vistos:
            return

        vistos.add(chave_unica)
        candidatos.append(normalizado)

    # ---------- START ----------

    async def start(self):
        if self.arquivo_entrada:
            eans = self.ler_eans_arquivo(self.arquivo_entrada)
        else:
            eans = [self.ean]

        for ean in eans:
            ean = str(ean).strip()
            self.logger.info(
                "Busca ean | valor=%s | URL=%s",
                ean,
                self.build_search_url(ean),
            )
            yield Request(
                url=self.build_search_url(ean),
                callback=self.parse_busca,
                dont_filter=True,
                meta={"ean_atual": ean},
            )

    # ---------- PARSE BUSCA ----------

    def parse_busca(self, response):
        ean = re.sub(r"\D", "", response.meta["ean_atual"])

        self.logger.info(
            "TIPO RESPONSE=%s | STATUS=%s | URL=%s",
            response.__class__.__name__,
            response.status,
            response.url,
        )

        candidatos = []
        vistos = set()

        # JSON-LD
        for bloco in response.css('script[type="application/ld+json"]::text').getall():
            try:
                data = json.loads(bloco)
            except Exception:
                continue

            estruturas = data if isinstance(data, list) else [data]

            for obj in estruturas:
                for d in self.walk_dicts(obj):
                    texto = " ".join(
                        str(v) for v in d.values() if isinstance(v, (str, int, float))
                    )
                    texto_norm = re.sub(r"\D", "", texto)

                    if ean and ean not in texto_norm:
                        continue

                    nome = d.get("name")
                    marca = d.get("brand")
                    if isinstance(marca, dict):
                        marca = marca.get("name")

                    offers = d.get("offers") or {}
                    if isinstance(offers, list):
                        offers = offers[0] if offers else {}

                    preco = (
                        d.get("price")
                        or offers.get("price")
                        or offers.get("lowPrice")
                        or offers.get("highPrice")
                    )

                    link = d.get("url") or response.url

                    self.add_candidate(
                        candidatos,
                        vistos,
                        {
                            "ean": ean,
                            "nome": nome,
                            "marca": marca,
                            "preco": preco,
                            "link": response.urljoin(link) if link else response.url,
                        },
                    )

        # scripts genéricos
        scripts = response.css("script::text").getall()
        for script in scripts:
            if ean and ean not in re.sub(r"\D", "", script):
                continue

            blocos = re.findall(r"\{.*?\}", script, flags=re.S)
            for trecho in blocos:
                trecho_lower = trecho.lower()
                if "price" not in trecho_lower and "product" not in trecho_lower:
                    continue

                try:
                    data = json.loads(trecho)
                except Exception:
                    continue

                for d in self.walk_dicts(data):
                    texto = " ".join(
                        str(v) for v in d.values() if isinstance(v, (str, int, float))
                    )
                    texto_norm = re.sub(r"\D", "", texto)

                    if ean and ean not in texto_norm:
                        continue

                    nome = (
                        d.get("productName")
                        or d.get("name")
                        or d.get("nombre")
                        or d.get("title")
                    )
                    marca = d.get("brand") or d.get("brandName") or d.get("marca")
                    preco = (
                        d.get("price")
                        or d.get("bestPrice")
                        or d.get("sellingPrice")
                        or d.get("precio")
                        or d.get("lowPrice")
                    )
                    link = (
                        d.get("link")
                        or d.get("url")
                        or d.get("productUrl")
                        or d.get("href")
                        or response.url
                    )

                    self.add_candidate(
                        candidatos,
                        vistos,
                        {
                            "ean": ean,
                            "nome": nome,
                            "marca": marca,
                            "preco": preco,
                            "link": response.urljoin(link) if link else response.url,
                        },
                    )

        # links de produto (PDP)
        for href in response.css('a[href*="/p"]::attr(href)').getall():
            href_abs = response.urljoin(href)
            if ean:
                self.add_candidate(
                    candidatos,
                    vistos,
                    {
                        "ean": ean,
                        "nome": None,
                        "marca": None,
                        "preco": None,
                        "link": href_abs,
                    },
                )

        # fallback na página de busca
        nome = (
            response.css("h1::text").get()
            or response.css('[class*="productName"]::text').get()
            or response.css("title::text").get()
        )
        marca = (
            response.css('[class*="productBrand"]::text').get()
            or response.css('[class*="brand"]::text').get()
        )
        preco = (
            response.css('[class*="sellingPrice"]::text').get()
            or response.css('[class*="price"]::text').get()
        )

        self.add_candidate(
            candidatos,
            vistos,
            {
                "ean": ean,
                "nome": nome,
                "marca": marca,
                "preco": preco,
                "link": response.url,
            },
        )

        if not candidatos:
            self.logger.warning(
                "Nenhum produto encontrado para EAN=%s | URL=%s",
                ean,
                response.url,
            )
            yield {
                "ean": ean,
                "nome": None,
                "marca": None,
                "preco": None,
                "loja": "carrefour_ar",
                "link": response.url,
            }
            return

        melhores = sorted(candidatos, key=self.item_score, reverse=True)
        melhor = melhores[0]

        self.logger.info(
            "EAN=%s | candidatos=%d | selecionado=%s | preco=%s | link=%s",
            ean,
            len(candidatos),
            melhor.get("nome"),
            melhor.get("preco"),
            melhor.get("link"),
        )

        link_melhor = melhor.get("link") or ""
        if "/p" in link_melhor and link_melhor.rstrip("/") != response.url.rstrip("/"):
            yield Request(
                url=link_melhor,
                callback=self.parse_produto,
                dont_filter=True,
                meta={"ean_atual": ean, "candidato_busca": melhor},
            )
            return

        yield melhor

    # ---------- PARSE PDP ----------

    def parse_produto(self, response):
        ean = re.sub(r"\D", "", response.meta["ean_atual"])
        base = response.meta.get("candidato_busca") or {}

        nome = (
            response.css("h1::text").get()
            or response.css('[class*="productName"]::text').get()
            or base.get("nome")
        )

        marca = (
            response.css('[class*="productBrand"]::text').get()
            or response.css('[class*="brand"]::text').get()
            or base.get("marca")
        )

        preco = (
            response.css('[class*="sellingPrice"]::text').get()
            or response.css('[class*="price"]::text').get()
            or base.get("preco")
        )

        for bloco in response.css('script[type="application/ld+json"]::text').getall():
            try:
                data = json.loads(bloco)
            except Exception:
                continue

            estruturas = data if isinstance(data, list) else [data]
            for obj in estruturas:
                for d in self.walk_dicts(obj):
                    nome = nome or d.get("name")

                    brand = d.get("brand")
                    if isinstance(brand, dict):
                        brand = brand.get("name")
                    marca = marca or brand

                    offers = d.get("offers") or {}
                    if isinstance(offers, list):
                        offers = offers[0] if offers else {}

                    preco = (
                        preco
                        or d.get("price")
                        or offers.get("price")
                        or offers.get("lowPrice")
                        or offers.get("highPrice")
                    )

        item = {
            "ean": ean,
            "nome": self.clean_nome(nome, ean) or base.get("nome"),
            "marca": self.normalize_text(marca),
            "preco": self.parse_price(preco),
            "loja": "carrefour_ar",
            "link": response.url,
        }

        self.logger.info(
            "PDP FINAL | EAN=%s | nome=%s | preco=%s | link=%s",
            item["ean"],
            item["nome"],
            item["preco"],
            item["link"],
        )

        yield item