import base64
import csv
import re
from pathlib import Path
from urllib.parse import quote

from scrapy import Spider, Request, Selector

try:
    import openpyxl
except ImportError:
    openpyxl = None


SITES = {
    "carrefour_br": {
        "base_url_busca": "https://mercado.carrefour.com.br/busca/{query}",
        "allowed_domains": [
            "mercado.carrefour.com.br",
            "carrefour.com.br",
            "www.carrefour.com.br",
        ],
    },
    "carrefour_ar": {
        "base_url_busca": "https://www.carrefour.com.ar/?keyword={query}",
        "allowed_domains": [
            "www.carrefour.com.ar",
            "carrefour.com.ar",
        ],
    },
    "jumbo_ar": {
        "base_url_busca": "https://www.jumbo.com.ar/busqueda?q={query}",
        "allowed_domains": [
            "www.jumbo.com.ar",
            "jumbo.com.ar",
        ],
    },
    "masonline_ar": {
        "base_url_busca": "https://www.masonline.com.ar/busqueda?q={query}",
        "allowed_domains": [
            "www.masonline.com.ar",
            "masonline.com.ar",
        ],
    },
}


class JumboPrecoSpider(Spider):
    name = "jumbo_preco"

    custom_settings = {
        "ZYTE_API_TRANSPARENT_MODE": True,
        "CONCURRENT_REQUESTS_PER_DOMAIN": 4,
        "DOWNLOAD_DELAY": 0.2,
        "LOG_LEVEL": "INFO",
    }

    def __init__(self, ean=None, arquivo_entrada=None, loja="jumbo_ar", *args, **kwargs):
        super().__init__(*args, **kwargs)

        loja = (loja or "jumbo_ar").lower()

        if loja not in SITES:
            raise ValueError(f"Loja '{loja}' não suportada. Use: {list(SITES.keys())}")

        if not ean and not arquivo_entrada:
            raise ValueError(
                "Passe ean ou arquivo_entrada. "
                "Ex.: -a ean=789... ou -a arquivo_entrada=seus_eans.csv/.xlsx"
            )

        self.ean = ean
        self.arquivo_entrada = arquivo_entrada
        self.loja = loja
        self.site_cfg = SITES[loja]
        self.allowed_domains = self.site_cfg["allowed_domains"]

        self.prints_dir = Path("prints")
        self.prints_dir.mkdir(exist_ok=True)

    def _ler_eans_csv(self, caminho: Path):
        eans = []

        with caminho.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)

            if not reader.fieldnames:
                raise ValueError("CSV sem cabeçalho.")

            nomes = {c.lower().strip(): c for c in reader.fieldnames}

            coluna_ean = (
                nomes.get("ean")
                or nomes.get("código ean")
                or nomes.get("codigo ean")
                or nomes.get("codigo_ean")
                or nomes.get("codigoean")
                or nomes.get("ean 13")
                or nomes.get("cod ean")
                or nomes.get("cod_ean")
            )

            if not coluna_ean:
                raise ValueError(
                    f"Não encontrei coluna de EAN no CSV. Cabeçalho: {reader.fieldnames}"
                )

            for row in reader:
                valor = (row.get(coluna_ean) or "").strip()
                if valor:
                    eans.append(valor)

        return list(dict.fromkeys(eans))

    def _ler_eans_xlsx(self, caminho: Path):
        if openpyxl is None:
            raise RuntimeError("openpyxl não está instalado. Rode: pip install openpyxl")

        wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Planilha vazia.")

        header_row_idx = None
        for i, row in enumerate(rows):
            if row is None:
                continue
            valores = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if valores:
                header_row_idx = i
                break

        if header_row_idx is None:
            raise ValueError("Planilha sem dados.")

        header = [str(h).strip() if h is not None else "" for h in rows[header_row_idx]]
        header_normalizado = [h.strip().lower() for h in header]
        nomes = {c: idx for idx, c in enumerate(header_normalizado)}

        idx_ean = (
            nomes.get("ean")
            or nomes.get("código ean")
            or nomes.get("codigo ean")
            or nomes.get("codigo_ean")
            or nomes.get("codigoean")
            or nomes.get("ean 13")
            or nomes.get("cod ean")
            or nomes.get("cod_ean")
        )

        if idx_ean is None:
            for nome_coluna, idx in nomes.items():
                if "ean" in nome_coluna:
                    idx_ean = idx
                    break

        if idx_ean is None:
            raise ValueError(
                "Não encontrei coluna de EAN no XLSX. "
                f"Cabeçalho original: {header} | Cabeçalho normalizado: {header_normalizado}"
            )

        eans = []
        for row in rows[header_row_idx + 1:]:
            if row is None:
                continue
            if all(c is None for c in row):
                continue
            if idx_ean >= len(row):
                continue

            valor = row[idx_ean]
            if valor is None:
                continue

            valor_str = str(valor).strip()
            if valor_str:
                eans.append(valor_str)

        return list(dict.fromkeys(eans))

    def _ler_eans_arquivo(self, caminho_str: str):
        caminho = Path(caminho_str)

        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo de entrada não encontrado: {caminho}")

        sufixo = caminho.suffix.lower()

        if sufixo == ".csv":
            return self._ler_eans_csv(caminho)
        elif sufixo == ".xlsx":
            return self._ler_eans_xlsx(caminho)
        else:
            raise ValueError(
                f"Extensão não suportada: {sufixo}. Use .csv ou .xlsx para arquivo_entrada."
            )

    def start_requests(self):
        if self.arquivo_entrada:
            eans = self._ler_eans_arquivo(self.arquivo_entrada)
            total_lidos = len(eans)
            eans = eans[:50]

            self.logger.info(
                "Lidos %d EAN(s) do arquivo %s | processando apenas os %d primeiros",
                total_lidos,
                self.arquivo_entrada,
                len(eans),
            )

            for ean in eans:
                query_enc = quote(str(ean))
                url = self.site_cfg["base_url_busca"].format(query=query_enc)

                yield Request(
                    url=url,
                    callback=self.parse_search,
                    dont_filter=True,
                    meta={
                        "ean_atual": str(ean),
                        "zyte_api_automap": {
                            "browserHtml": True,
                        },
                    },
                )
            return

        query_enc = quote(str(self.ean))
        url = self.site_cfg["base_url_busca"].format(query=query_enc)

        yield Request(
            url=url,
            callback=self.parse_search,
            dont_filter=True,
            meta={
                "ean_atual": str(self.ean),
                "zyte_api_automap": {
                    "browserHtml": True,
                },
            },
        )

    def slugify(self, texto: str) -> str:
        texto = (texto or "").strip().lower()
        texto = re.sub(r"[^\w\s-]", "", texto, flags=re.UNICODE)
        texto = re.sub(r"[-\s]+", "-", texto)
        return (texto[:120] or "item").strip("-") or "item"

    def extrair_preco(self, seletor):
        if hasattr(seletor, "getall"):
            textos = seletor.getall()
        else:
            textos = seletor.css("::text").getall()

        bloco_texto = " ".join(t.strip() for t in textos if t and t.strip())

        padroes = [
            r"(R\$\s*\d{1,3}(?:\.\d{3})*,\d{2})",
            r"(\$\s*\d{1,3}(?:\.\d{3})*,\d{2})",
            r"(\$\s*\d{1,3}(?:\.\d{3})+)",
            r"(\$\s*\d+,\d{2})",
            r"(\$\s*\d+)",
        ]

        for padrao in padroes:
            match = re.search(padrao, bloco_texto, flags=re.IGNORECASE)
            if match:
                return match.group(1)

        return None

    def extrair_desconto(self, seletor):
        if hasattr(seletor, "getall"):
            textos = seletor.getall()
        else:
            textos = seletor.css("::text").getall()

        bloco_texto = " ".join(t.strip() for t in textos if t and t.strip())

        m = re.search(r"(\d{1,3})\s*%", bloco_texto)
        if m:
            return m.group(1)
        return None

    def extrair_marca(self, nome_limpo: str):
        palavras = (nome_limpo or "").split()
        if not palavras:
            return None
        return palavras[0]

    def salvar_screenshot(self, item, response):
        raw = getattr(response, "raw_api_response", None) or {}
        screenshot_b64 = raw.get("screenshot")

        if not screenshot_b64:
            return None

        nome_arquivo = f"{self.slugify(item.get('nome') or 'item')}.png"
        caminho_arquivo = self.prints_dir / nome_arquivo

        with open(caminho_arquivo, "wb") as f:
            f.write(base64.b64decode(screenshot_b64))

        return str(caminho_arquivo.resolve())

    def _get_html_selector(self, response):
        raw = getattr(response, "raw_api_response", None) or {}
        browser_html = raw.get("browserHtml")
        if browser_html:
            return Selector(text=browser_html)
        return response

    def obter_produtos_da_pagina(self, response):
        response_sel = self._get_html_selector(response)

        seletores = [
            "[class*='vtex-product-summary']",
            "section[data-testid*='product-summary']",
            "article[data-testid*='product-summary']",
            "div[data-testid*='product-summary']",
            "article",
            "[class*='product']",
            "a[href*='/p']",
        ]

        vistos = []
        chaves_vistas = set()

        for seletor in seletores:
            for node in response_sel.css(seletor):
                href = node.css("a::attr(href), ::attr(href)").get()
                textos = node.css("::text").getall()
                texto_base = "".join(textos[:3]).strip() if textos else ""
                chave = f"{href or ''}|{texto_base}"

                if chave not in chaves_vistas:
                    chaves_vistas.add(chave)
                    vistos.append(node)

        return vistos

    def extrair_nome_produto_lista(self, produto):
        seletores_nome = [
            "h1::text",
            "h2::text",
            "h3::text",
            "h4::text",
            "a::text",
            "[class*='name']::text",
            "[class*='title']::text",
            "[class*='brand']::text",
            "[data-testid*='name']::text",
        ]

        for seletor in seletores_nome:
            nome = produto.css(seletor).get()
            if nome and nome.strip():
                nome_limpo = " ".join(nome.split())
                if len(nome_limpo) > 2:
                    return nome_limpo

        textos = produto.css("::text").getall()
        textos_limpos = [t.strip() for t in textos if t and t.strip()]
        for txt in textos_limpos:
            if len(txt) > 2 and txt.lower() not in {"patrocinado", "sponsored"}:
                return " ".join(txt.split())

        return None

    def extrair_link_produto_lista(self, produto, response):
        response_sel = self._get_html_selector(response)
        seletores_link = [
            "a[href*='/producto/']::attr(href)",
            "a[href*='/p/']::attr(href)",
            "a[href*='/p']::attr(href)",
            "a::attr(href)",
            "::attr(href)",
        ]

        for seletor in seletores_link:
            link = produto.css(seletor).get()
            if link:
                return response_sel.urljoin(link)

        return None

    def parse_search(self, response):
        ean_atual = response.meta.get("ean_atual") or self.ean

        self.logger.info("Loja: %s | EAN: %s | URL: %s", self.loja, ean_atual, response.url)

        produtos = self.obter_produtos_da_pagina(response)
        self.logger.info("Produtos encontrados na página: %s", len(produtos))

        for produto in produtos:
            nome_limpo = self.extrair_nome_produto_lista(produto)
            if not nome_limpo:
                continue

            preco_lista = self.extrair_preco(produto)
            desconto_percentual_lista = self.extrair_desconto(produto)
            marca_lista = self.extrair_marca(nome_limpo)
            link_absoluto = self.extrair_link_produto_lista(produto, response)

            if not link_absoluto:
                continue

            item_base = {
                "loja": self.loja,
                "ean": ean_atual,
                "nome": nome_limpo,
                "marca": marca_lista,
                "preco": preco_lista,
                "desconto_percentual": desconto_percentual_lista,
                "print_tela_path": None,
                "link": link_absoluto,
            }

            yield Request(
                url=link_absoluto,
                callback=self.parse_produto,
                dont_filter=True,
                meta={
                    "item_base": item_base,
                    "zyte_api_automap": {
                        "browserHtml": True,
                        "screenshot": True,
                    },
                },
            )

    def parse_produto(self, response):
        response_sel = self._get_html_selector(response)
        item = response.meta["item_base"].copy()

        nome_pagina = response_sel.css(
            "h1::text, [class*='product'] h1::text, [class*='title']::text"
        ).get()

        if nome_pagina:
            item["nome"] = " ".join(nome_pagina.split())

        textos = response_sel.css("body ::text").getall()
        bloco_texto = " ".join(t.strip() for t in textos if t and t.strip())

        if not item.get("preco"):
            padroes_preco = [
                r"(R\$\s*\d{1,3}(?:\.\d{3})*,\d{2})",
                r"(\$\s*\d{1,3}(?:\.\d{3})*,\d{2})",
                r"(\$\s*\d{1,3}(?:\.\d{3})+)",
                r"(\$\s*\d+,\d{2})",
                r"(\$\s*\d+)",
            ]
            for padrao in padroes_preco:
                m_preco = re.search(padrao, bloco_texto)
                if m_preco:
                    item["preco"] = m_preco.group(1)
                    break

        if not item.get("desconto_percentual"):
            m_desc = re.search(r"(\d{1,3})\s*%", bloco_texto)
            if m_desc:
                item["desconto_percentual"] = m_desc.group(1)

        if not item.get("marca") and item.get("nome"):
            item["marca"] = self.extrair_marca(item["nome"])

        try:
            item["print_tela_path"] = self.salvar_screenshot(item, response)
        except Exception:
            item["print_tela_path"] = None

        yield item