import base64
import json
import re

import scrapy
from openpyxl import load_workbook


class JumboSearchSpider(scrapy.Spider):
    name = "jumbo_search"
    allowed_domains = ["jumbo.com.ar"]

    custom_settings = {
        "LOG_LEVEL": "INFO",
        "DOWNLOAD_TIMEOUT": 60,
        "DOWNLOAD_DELAY": 0.8,
        "CONCURRENT_REQUESTS_PER_DOMAIN": 2,
        "FEED_EXPORT_ENCODING": "utf-8",
        "FEED_EXPORT_FIELDS": [
            "ean",
            "sku",
            "nome",
            "marca",
            "precoDe",
            "precoPor",
            "porcentagem",
            "oferta",
            "loja",
            "link",
        ],
    }

    SEARCH_URL = "https://www.jumbo.com.ar/api/catalog_system/pub/products/search"

    def __init__(self, ean=None, ean_file=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.eans = []

        if ean:
            self.eans = [str(ean).strip()]
        elif ean_file:
            self.eans = self.load_eans_from_excel(ean_file)

        self.logger.info(f"Total de EANs carregados: {len(self.eans)}")

    def load_eans_from_excel(self, file_path):
        eans = []

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active

            headers = next(
                ws.iter_rows(min_row=1, max_row=1, values_only=True),
                None,
            )

            if not headers:
                self.logger.warning("Arquivo Excel sem cabeçalho.")
                wb.close()
                return eans

            ean_col_idx = None
            for idx, header in enumerate(headers):
                if header and str(header).strip().upper() == "EAN":
                    ean_col_idx = idx
                    break

            if ean_col_idx is None:
                self.logger.warning("Coluna 'EAN' não encontrada no arquivo.")
                wb.close()
                return eans

            for row in ws.iter_rows(min_row=2, values_only=True):
                valor = row[ean_col_idx] if len(row) > ean_col_idx else None
                if valor is None:
                    continue

                ean = str(valor).strip()
                if ean:
                    eans.append(ean)

            wb.close()

        except Exception as e:
            self.logger.error(f"Erro lendo arquivo Excel de EANs: {e}")

        return eans

    def start_requests(self):
        for ean in self.eans:
            url = f"{self.SEARCH_URL}?fq=alternateIds_Ean:{ean}"
            yield scrapy.Request(
                url=url,
                callback=self.parse_search,
                meta={"ean": ean},
            )

    def parse_search(self, response):
        ean = response.meta["ean"]

        try:
            data = json.loads(response.text)
        except Exception:
            self.logger.warning(f"JSON inválido na busca para EAN {ean}")
            return

        if not data:
            self.logger.info(f"Nenhum produto encontrado para EAN {ean}")
            return

        product = data[0]
        nome = product.get("productName")
        marca = product.get("brand")
        link_text = product.get("linkText")
        items = product.get("items") or []

        if not items:
            self.logger.info(f"Produto sem items para EAN {ean}")
            return

        item0 = items[0]
        sku_api = item0.get("itemId")
        eans_api = item0.get("ean") or ean

        sellers = item0.get("sellers") or []
        seller0 = sellers[0] if sellers else {}
        commertial = seller0.get("commertialOffer") or {}

        fallback_price = self.normalize_price(commertial.get("Price"))
        fallback_list_price = self.normalize_price(commertial.get("ListPrice"))

        self.logger.info(
            f"EAN {ean} | SKU API={sku_api} | "
            f"Price bruto={commertial.get('Price')} | "
            f"ListPrice bruto={commertial.get('ListPrice')} | "
            f"Price={fallback_price} | ListPrice={fallback_list_price}"
        )

        product_url = f"https://www.jumbo.com.ar/{link_text}/p"

        yield scrapy.Request(
            url=product_url,
            callback=self.parse_product,
            meta={
                "ean": str(eans_api or ean),
                "sku_api": str(sku_api) if sku_api is not None else None,
                "nome_api": nome,
                "marca_api": marca,
                "link": product_url,
                "fallback_price": fallback_price,
                "fallback_list_price": fallback_list_price,
                "zyte_api_automap": {
                    "browserHtml": True,
                    "networkCapture": [
                        {
                            "filterType": "url",
                            "matchType": "contains",
                            "value": "/_v/search-promotions",
                            "httpResponseBody": True,
                        }
                    ],
                },
            },
        )

    def parse_product(self, response):
        ean = response.meta.get("ean")
        sku_api = response.meta.get("sku_api")
        nome_api = response.meta.get("nome_api")
        marca_api = response.meta.get("marca_api")
        link = response.meta.get("link")
        fallback_price = response.meta.get("fallback_price")
        fallback_list_price = response.meta.get("fallback_list_price")

        sku = self.extract_sku(response.text) or sku_api
        nome = self.extract_nome(response.text) or nome_api
        marca = self.extract_marca(response.text) or marca_api

        self.logger.info(
            f"PDP EAN {ean} | SKU visual={sku} | SKU API={sku_api}"
        )

        promo_percent, eff_raw = self.extract_discount_from_promotions(
            response,
            sku_visual=str(sku) if sku is not None else None,
            sku_api=str(sku_api) if sku_api is not None else None,
        )

        item = self.mount_final_item(
            ean=ean,
            sku=sku,
            nome=nome,
            marca=marca,
            link=link,
            promo_percent=promo_percent,
            effective_discount_raw=eff_raw,
            fallback_price=fallback_price,
            fallback_list_price=fallback_list_price,
        )

        self.logger.info(f"Item final EAN {ean}: {item}")
        yield item

    def extract_discount_from_promotions(
        self,
        response,
        sku_visual=None,
        sku_api=None,
    ):
        raw_api_response = getattr(response, "raw_api_response", None) or {}
        captures = raw_api_response.get("networkCapture") or []

        self.logger.info("NETWORK_CAPTURE SIZE: %s", len(captures))

        sku_candidates = [str(x) for x in (sku_visual, sku_api) if x]
        self.logger.info("SKU candidates promo: %s", sku_candidates)

        for cap in captures:
            url = cap.get("url") or ""
            self.logger.info("CAPTURE URL: %s", url)

            if "/_v/search-promotions" not in url:
                continue

            body_b64 = cap.get("httpResponseBody")
            if not body_b64:
                self.logger.info("CAPTURE sem httpResponseBody")
                continue

            try:
                raw = base64.b64decode(body_b64).decode(
                    "utf-8",
                    errors="ignore",
                )
                data = json.loads(raw)
            except Exception as e:
                self.logger.info("Erro decodificando promo body: %s", e)
                continue

            generic_promos = (
                data.get("promotions", {})
                .get("generic", {})
                .get("promotions", {})
            )

            self.logger.info(
                "Promo SKUs disponíveis: %s",
                list(generic_promos.keys())[:20],
            )

            for sku in sku_candidates:
                promo = generic_promos.get(sku)
                if not promo:
                    continue

                effective_discount = promo.get("effectiveDiscount")
                self.logger.info(
                    "Promo encontrada para SKU %s | effectiveDiscount=%s",
                    sku,
                    effective_discount,
                )

                if effective_discount is None:
                    continue

                try:
                    eff_float = float(str(effective_discount))
                    pct = int(round(eff_float * 100))
                    return pct, eff_float
                except Exception as e:
                    self.logger.info(
                        "Erro convertendo effectiveDiscount: %s",
                        e,
                    )
                    continue

        return None, None

    def extract_sku(self, html):
        patterns = [
            r"SKU\s*:?\s*(\d{5,20})",
            r'"sku"\s*:\s*"(\d+)"',
            r'"itemId"\s*:\s*"(\d+)"',
        ]

        for pattern in patterns:
            match = re.search(pattern, html, flags=re.IGNORECASE)
            if match:
                return match.group(1)

        return None

    def extract_nome(self, html):
        patterns = [
            r"<title>\s*(.*?)\s*-\s*Jumbo\s*</title>",
            r'"productName"\s*:\s*"([^"]+)"',
            r"<h1[^>]*>\s*(.*?)\s*</h1>",
        ]

        for pattern in patterns:
            match = re.search(
                pattern,
                html,
                flags=re.IGNORECASE | re.DOTALL,
            )
            if match:
                return self.clean_spaces(
                    self.strip_tags(match.group(1))
                )

        return None

    def extract_marca(self, html):
        patterns = [
            r'"brand"\s*:\s*"([^"]+)"',
            r'"Marca"\s*:\s*"([^"]+)"',
        ]

        for pattern in patterns:
            match = re.search(pattern, html, flags=re.IGNORECASE)
            if match:
                return self.clean_spaces(match.group(1))

        match = re.search(
            r"NIVEA|DERMAGLOS|DOVE|REXONA",
            html,
            flags=re.IGNORECASE,
        )
        if match:
            return match.group(0).upper()

        return None

    def mount_final_item(
        self,
        ean,
        sku,
        nome,
        marca,
        link,
        promo_percent,
        effective_discount_raw,
        fallback_price,
        fallback_list_price,
    ):
        preco_de = None
        preco_por = None
        porcentagem = None
        oferta = None

        if fallback_price is not None:
            preco_de = fallback_price

        if promo_percent is not None:
            porcentagem = promo_percent
            oferta = "X"

        if preco_de is not None and porcentagem is not None:
            try:
                preco_por = round(
                    preco_de * (1 - (porcentagem / 100.0)),
                    2,
                )
            except Exception:
                preco_por = None

        return {
            "ean": str(ean) if ean is not None else None,
            "sku": str(sku) if sku is not None else None,
            "nome": nome,
            "marca": marca,
            "precoDe": preco_de,
            "precoPor": preco_por,
            "effective_discount_raw": effective_discount_raw,
            "porcentagem": porcentagem,
            "oferta": oferta,
            "loja": "Jumbo",
            "link": link,
        }

    def normalize_price(self, value):
        if value is None or value == "":
            return None

        if isinstance(value, (int, float)):
            return float(value)

        s = str(value).strip()
        s = s.replace("$", "").replace("\xa0", "").replace(" ", "")
        s = re.sub(r"[^\d,.\-]", "", s)

        if not s:
            return None

        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            parts = s.split(",")
            if len(parts[-1]) in (1, 2):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "." in s:
            parts = s.split(".")
            if len(parts[-1]) in (1, 2):
                pass
            else:
                s = s.replace(".", "")

        try:
            return float(s)
        except Exception:
            return None

    def strip_tags(self, value):
        return re.sub(r"<[^>]+>", " ", value or "")

    def clean_spaces(self, value):
        return re.sub(r"\s+", " ", (value or "")).strip()