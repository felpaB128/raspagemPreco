import csv
import json
import re
import unicodedata
from pathlib import Path
from urllib.parse import quote

from scrapy import Spider, Request

try:
    import openpyxl
except ImportError:
    openpyxl = None


class MasonlineSpider(Spider):
    name = "masonline"
    allowed_domains = ["masonline.com.ar"]

    custom_settings = {
        "LOG_LEVEL": "INFO",
        "DOWNLOAD_DELAY": 0.25,
        "CONCURRENT_REQUESTS_PER_DOMAIN": 4,
        "DOWNLOAD_TIMEOUT": 60,
        "FEED_EXPORT_ENCODING": "utf-8",
    }

    def __init__(self, arquivo_entrada=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if not arquivo_entrada:
            raise ValueError("Passe -a arquivo_entrada=seu_arquivo.xlsx")
        self.arquivo_entrada = arquivo_entrada

    def _resolver_caminho_arquivo(self, caminho_str: str) -> Path:
        caminho = Path(caminho_str).expanduser()
        if caminho.is_absolute():
            return caminho
        return Path.cwd() / caminho

    def _normalize_colname(self, s):
        if s is None:
            return ""
        s = str(s).strip().lower()
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        s = re.sub(r"\s+", " ", s)
        return s

    def _clean_value(self, s):
        if s is None:
            return None
        s = str(s).strip()
        if not s or s in {"-", "--", "nan", "None", "none"}:
            return None
        return s

    def _norm_text(self, s):
        if not s:
            return ""
        s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
        s = s.lower().strip()
        s = re.sub(r"[^a-z0-9\s]", " ", s)
        s = re.sub(r"\s+", " ", s)
        return s.strip()

    def _norm_digits(self, s):
        if s is None:
            return ""
        return re.sub(r"\D", "", str(s))

    def _is_meaningful_text(self, s):
        s = self._clean_value(s)
        if not s:
            return False
        n = self._norm_text(s)
        return bool(n and n not in {"-", "na", "n a", "null"})

    def _normalize_price(self, value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return f"{float(value):.2f}"
        txt = str(value).strip()
        if not txt:
            return None
        txt = txt.replace("\xa0", " ").replace("$", "").strip()
        txt = re.sub(r"[^\d,\.]", "", txt)
        if not txt:
            return None
        if "," in txt and "." in txt:
            txt = txt.replace(".", "").replace(",", ".")
        elif "," in txt:
            txt = txt.replace(".", "").replace(",", ".")
        else:
            txt = txt.replace(",", "")
        try:
            return f"{float(txt):.2f}"
        except Exception:
            return None

    def montar_url_busca_runtime(self, termo: str) -> str:
        q = quote(str(termo).strip(), safe="")
        return (
            f"https://www.masonline.com.ar/{q}"
            f"?map=ft&_q={q}&__pickRuntime=query,queryData&__device=desktop"
        )

    def _read_xlsx(self, caminho: Path):
        if openpyxl is None:
            raise RuntimeError("Instale openpyxl: pip install openpyxl")

        wb = openpyxl.load_workbook(str(caminho), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Planilha vazia")

        header = [self._normalize_colname(x) for x in rows[0]]
        idx_map = {}

        for i, col in enumerate(header):
            if col == "ean":
                idx_map["ean"] = i
            elif col == "nombre":
                idx_map["nombre"] = i
            elif col == "marca":
                idx_map["marca"] = i
            elif col == "sku":
                idx_map["sku"] = i

        itens = []
        for row in rows[1:]:
            if not row:
                continue

            item = {
                "ean": self._clean_value(row[idx_map["ean"]]) if "ean" in idx_map and idx_map["ean"] < len(row) else None,
                "nombre": self._clean_value(row[idx_map["nombre"]]) if "nombre" in idx_map and idx_map["nombre"] < len(row) else None,
                "marca": self._clean_value(row[idx_map["marca"]]) if "marca" in idx_map and idx_map["marca"] < len(row) else None,
                "sku": self._clean_value(row[idx_map["sku"]]) if "sku" in idx_map and idx_map["sku"] < len(row) else None,
            }

            if any(item.values()):
                itens.append(item)

        return itens

    def _read_csv(self, caminho: Path):
        itens = []
        with caminho.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            header = {self._normalize_colname(c): c for c in reader.fieldnames or []}

            for row in reader:
                item = {
                    "ean": self._clean_value(row.get(header.get("ean"))),
                    "nombre": self._clean_value(row.get(header.get("nombre"))),
                    "marca": self._clean_value(row.get(header.get("marca"))),
                    "sku": self._clean_value(row.get(header.get("sku"))),
                }
                if any(item.values()):
                    itens.append(item)

        return itens

    def _ler_arquivo(self):
        caminho = self._resolver_caminho_arquivo(self.arquivo_entrada)

        if not caminho.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

        if caminho.suffix.lower() == ".xlsx":
            return self._read_xlsx(caminho)

        if caminho.suffix.lower() == ".csv":
            return self._read_csv(caminho)

        raise ValueError("Use arquivo .xlsx ou .csv")

    def _contains_whole_text(self, needle, haystack):
        n = self._norm_text(needle)
        h = self._norm_text(haystack)
        if not n or not h:
            return False
        return n in h

    def _token_overlap_ratio(self, a, b):
        ta = set(self._norm_text(a).split())
        tb = set(self._norm_text(b).split())
        if not ta or not tb:
            return 0.0
        inter = ta.intersection(tb)
        return len(inter) / max(1, len(ta))

    def _compute_match(self, origem, candidato):
        match = {"ean": False, "sku": False, "nombre": False, "marca": False}

        origem_ean = self._norm_digits(origem.get("ean"))
        origem_sku = self._norm_digits(origem.get("sku"))
        origem_nombre = self._clean_value(origem.get("nombre"))
        origem_marca = self._clean_value(origem.get("marca"))

        cand_ean = self._norm_digits(candidato.get("ean"))
        cand_sku = self._norm_digits(candidato.get("sku"))
        cand_nombre = self._clean_value(candidato.get("nombre"))
        cand_marca = self._clean_value(candidato.get("marca"))

        if origem_ean and cand_ean and origem_ean == cand_ean:
            match["ean"] = True

        if origem_sku and cand_sku and origem_sku == cand_sku:
            match["sku"] = True

        if self._is_meaningful_text(origem_nombre) and self._is_meaningful_text(cand_nombre):
            if self._contains_whole_text(origem_nombre, cand_nombre):
                match["nombre"] = True
            else:
                ratio = self._token_overlap_ratio(origem_nombre, cand_nombre)
                if ratio >= 0.6:
                    match["nombre"] = True

        if self._is_meaningful_text(origem_marca) and self._is_meaningful_text(cand_marca):
            om = self._norm_text(origem_marca)
            cm = self._norm_text(cand_marca)
            if om == cm or om in cm or cm in om:
                match["marca"] = True

        total = sum(1 for v in match.values() if v)
        score = (
            4 * int(match["ean"])
            + 3 * int(match["sku"])
            + 3 * int(match["nombre"])
            + 2 * int(match["marca"])
        )
        valido = total >= 2

        return match, total, score, valido

    def _build_output(self, loja, candidato):
        return {
            "loja": loja,
            "ean": candidato.get("ean"),
            "sku": candidato.get("sku"),
            "nome": candidato.get("nombre"),
            "marca": candidato.get("marca"),
            "precoDe": candidato.get("precoDe"),
            "precoPor": candidato.get("precoPor"),
            "oferta": candidato.get("oferta"),
            "print_tela_path": candidato.get("print_tela_path"),
            "link": candidato.get("link"),
        }

    def _extract_json_text(self, response):
        txt = response.text.strip()
        if not txt:
            return None

        start = txt.find("{")
        if start == -1:
            return None

        end = txt.rfind("}")
        if end == -1 or end <= start:
            return None

        return txt[start:end + 1]

    def _parse_querydata_payload(self, response):
        raw_json = self._extract_json_text(response)
        if not raw_json:
            return []

        try:
            payload = json.loads(raw_json)
        except Exception as e:
            self.logger.warning("Falha json.loads payload principal %s | erro=%r", response.url, e)
            return []

        query_data = payload.get("queryData") or []
        candidatos = []

        for block in query_data:
            data_str = block.get("data")
            if not data_str:
                continue

            try:
                data_obj = json.loads(data_str)
            except Exception:
                continue

            product_search = data_obj.get("productSearch") or {}
            products = product_search.get("products") or []

            for product in products:
                product_name = product.get("productName")
                brand = product.get("brand")
                link = product.get("link")
                items = product.get("items") or []

                for item in items:
                    item_id = item.get("itemId")
                    ean = item.get("ean")
                    sellers = item.get("sellers") or []
                    seller = sellers[0] if sellers else {}
                    offer = seller.get("commertialOffer") or {}

                    price = self._normalize_price(offer.get("Price"))
                    list_price = self._normalize_price(offer.get("ListPrice"))

                    precoDe = None
                    precoPor = None
                    oferta = None

                    try:
                        p = float(price) if price else None
                        lp = float(list_price) if list_price else None
                    except Exception:
                        p, lp = None, None

                    if p is not None and lp is not None:
                        if lp > p:
                            precoDe = f"{lp:.2f}"
                            precoPor = f"{p:.2f}"
                            oferta = "X"
                        else:
                            precoDe = f"{p:.2f}"
                            precoPor = None
                            oferta = None
                    elif p is not None:
                        precoDe = f"{p:.2f}"
                        precoPor = None
                        oferta = None
                    elif lp is not None:
                        precoDe = f"{lp:.2f}"
                        precoPor = None
                        oferta = None

                    candidatos.append({
                        "ean": ean,
                        "sku": item_id,
                        "nombre": product_name or item.get("nameComplete") or item.get("name"),
                        "marca": brand,
                        "precoDe": precoDe,
                        "precoPor": precoPor,
                        "oferta": oferta,
                        "link": response.urljoin(link) if link else response.url,
                        "print_tela_path": None,
                    })

        return candidatos

    def _montar_buscas_priorizadas(self, item):
        buscas = []

        if self._is_meaningful_text(item.get("nombre")):
            buscas.append(("nombre", item["nombre"]))

        if self._norm_digits(item.get("ean")):
            buscas.append(("ean", item["ean"]))

        if self._is_meaningful_text(item.get("marca")):
            buscas.append(("marca", item["marca"]))

        if self._norm_digits(item.get("sku")):
            buscas.append(("sku", item["sku"]))

        final = []
        vistos = set()

        for tipo, termo in buscas:
            chave = (tipo, self._norm_text(termo) if tipo in {"nombre", "marca"} else self._norm_digits(termo))
            if chave not in vistos:
                vistos.add(chave)
                final.append((tipo, termo))

        return final

    async def start(self):
        itens = self._ler_arquivo()
        self.logger.info("Itens lidos: %d", len(itens))

        for idx, item in enumerate(itens, start=1):
            buscas = self._montar_buscas_priorizadas(item)
            if not buscas:
                continue

            first_tipo, first_termo = buscas[0]

            yield Request(
                url=self.montar_url_busca_runtime(first_termo),
                callback=self.parse_search_runtime,
                dont_filter=True,
                meta={
                    "origem": item,
                    "item_index": idx,
                    "buscas": buscas,
                    "busca_pos": 0,
                    "tipo_busca": first_tipo,
                    "termo_buscado": first_termo,
                },
            )

    def parse_search_runtime(self, response):
        origem = response.meta["origem"]
        buscas = response.meta["buscas"]
        busca_pos = response.meta["busca_pos"]
        tipo_busca = response.meta["tipo_busca"]
        termo_buscado = response.meta["termo_buscado"]

        candidatos = self._parse_querydata_payload(response)

        melhor = None
        melhor_score = -1
        melhor_total = -1
        melhor_match = None

        for cand in candidatos:
            match, total, score, valido = self._compute_match(origem, cand)
            if valido and (score > melhor_score or (score == melhor_score and total > melhor_total)):
                melhor = cand
                melhor_score = score
                melhor_total = total
                melhor_match = match

        if melhor is not None:
            yield self._build_output(loja="masonline", candidato=melhor)
            return

        prox = busca_pos + 1
        if prox < len(buscas):
            prox_tipo, prox_termo = buscas[prox]
            yield Request(
                url=self.montar_url_busca_runtime(prox_termo),
                callback=self.parse_search_runtime,
                dont_filter=True,
                meta={
                    "origem": origem,
                    "buscas": buscas,
                    "busca_pos": prox,
                    "tipo_busca": prox_tipo,
                    "termo_buscado": prox_termo,
                },
            )
            return

        yield 